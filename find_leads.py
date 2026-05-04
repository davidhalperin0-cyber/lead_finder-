#!/usr/bin/env python3
"""
מנוע חיפוש לידים + עוזר הכנה לשיחה (אנושי בלבד)
===============================================
מחפש עסקים, מנתח אתר (כולל ניתוח AI אופציונלי), מחלץ קשר,
ומייצא אקסל + דשבורד HTML להכנה לשיחה — בלי שליחת הודעות אוטומטית.

שימוש:
    העתיקי .env.example ל-.env ומלאי OPENAI_API_KEY (או export בטרמינל)
    python find_leads.py --city "תל אביב" --type "מסעדה" --limit 30
    python find_leads.py ... --screenshots   # צילומים (Playwright)

תלויות: ראה requirements.txt
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

# ----------------------------------------------------------------------------
# תצורה
# ----------------------------------------------------------------------------

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
REQUEST_TIMEOUT = 12
HEADERS = {"User-Agent": USER_AGENT, "Accept-Language": "he,en;q=0.8"}


def _load_env_file() -> None:
    """טוען .env מתיקיית הפרויקט (ליד find_leads.py). אופציונלי אם python-dotenv מותקן."""
    try:
        from dotenv import load_dotenv
    except ImportError:
        return
    env_path = Path(__file__).resolve().parent / ".env"
    load_dotenv(env_path)


def _url_for_href(url: str) -> Optional[str]:
    """מנרמל URL לקישור בדפדפן / באקסל (טקסט בלבד לעיתים לא נפתח בלחיצה)."""
    if not url or not isinstance(url, str):
        return None
    u = url.strip()
    if not u:
        return None
    if u.startswith("//"):
        u = "https:" + u
    elif not re.match(r"^https?://", u, re.I):
        u = "https://" + u
    return u


def _tel_href(phone: str) -> Optional[str]:
    if not phone or not isinstance(phone, str):
        return None
    d = re.sub(r"\D", "", phone.strip())
    if len(d) < 9:
        return None
    if d.startswith("972"):
        return "tel:+" + d
    if d.startswith("0"):
        return "tel:+972" + d[1:]
    return "tel:+" + d


def _normalize_il_phone(phone: str) -> str:
    """מנקה טלפון ישראלי לפורמט אחיד +972XXXXXXXXX (או ריק)."""
    if not phone or not isinstance(phone, str):
        return ""
    d = re.sub(r"\D", "", phone)
    if len(d) < 9:
        return ""
    if d.startswith("00"):
        d = d[2:]
    if d.startswith("972"):
        return "+972" + d[3:]
    if d.startswith("0"):
        return "+972" + d[1:]
    # fallback
    return "+{}".format(d) if not d.startswith("+") else d


def _is_israeli_phone(phone: str) -> bool:
    """
    בודק אם טלפון נראה ישראלי (קידומת 972 או 0X תקינה).
    מטרה: לסנן עסקים שלא בישראל - למשל אם AI החזיר עסק עם טלפון אמריקאי (+1).
    """
    if not phone:
        return False
    d = re.sub(r"\D", "", phone)
    if d.startswith("00"):
        d = d[2:]
    # 972 בתחילת המספר - תקין (12 ספרות סה"כ)
    if d.startswith("972") and 11 <= len(d) <= 13:
        return True
    # 0 בתחילת המספר - תקין רק אם הספרה השנייה מהווה קידומת ישראלית
    # קידומות נייד: 050,051,052,053,054,055,058
    # קידומות נייח: 02,03,04,08,09
    if d.startswith("0") and 9 <= len(d) <= 10:
        if len(d) == 10 and d[1] == "5":  # נייד
            return True
        if len(d) == 9 and d[1] in ("2", "3", "4", "8", "9"):  # נייח
            return True
    return False


def _normalize_url_for_storage(url: str) -> str:
    """URL נקי לשמירה: מוריד query/fragment, מוריד www, מסיר / בסוף."""
    href = _url_for_href(url) or ""
    if not href:
        return ""
    try:
        p = urlparse(href)
        host = (p.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = p.path or ""
        if path != "/" and path.endswith("/"):
            path = path[:-1]
        clean = f"{p.scheme.lower()}://{host}{path}"
        return clean
    except Exception:
        return href.strip()


def _site_key_from_url(url: str) -> str:
    """מפתח דה-דופליקציה: host+path (ללא scheme)."""
    u = _normalize_url_for_storage(url)
    if not u:
        return ""
    try:
        p = urlparse(u)
        host = (p.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = (p.path or "").rstrip("/")
        return (host + path).strip()
    except Exception:
        return u.lower().strip()


# ----------------------------------------------------------------------------
# מודל נתונים
# ----------------------------------------------------------------------------

@dataclass
class Lead:
    business_name: str = ""
    website: str = ""
    final_url: str = ""
    score: int = 0
    grade: str = ""           # A=מעולה, B=טוב, C=בינוני, D=ישן, F=ישן מאוד
    issues: list = field(default_factory=list)
    email: str = ""
    phone: str = ""
    whatsapp: str = ""
    facebook: str = ""
    instagram: str = ""
    address: str = ""
    last_copyright: str = ""
    has_https: bool = False
    is_mobile_friendly: bool = False
    cms: str = ""             # WordPress / Wix / Custom / וכו'
    notes: str = ""
    error: str = ""
    # ניתוח AI — עוזר מחשבה לשיחה אנושית בלבד (לא שולח הודעות)
    ai_summary: str = ""
    main_problems: list = field(default_factory=list)
    ux_issues: list = field(default_factory=list)
    trust_issues: list = field(default_factory=list)
    conversion_issues: list = field(default_factory=list)
    best_talking_point: str = ""
    suggested_angle: str = ""
    priority_level: str = ""   # high / medium / low
    ai_notes: str = ""         # שגיאת API או "לא הופעל"
    screenshot_path: str = ""  # נתיב מקומי לתמונה (אופציונלי)
    search_city: str = ""
    search_business_type: str = ""
    opportunity_score: int = 0
    close_probability: int = 0
    strongest_problem: str = ""
    business_impact: str = ""
    opening_line: str = ""
    if_not_interested: str = ""
    what_to_offer: str = ""
    next_action: str = ""      # call | whatsapp | skip
    site_key: str = ""         # dedupe: host+path
    last_analyzed_at: str = "" # ISO string (UTC) או ריק
    match_score: int = 0       # 0-100, התאמה לתיאור החופשי שהמשתמש נתן
    match_reason: str = ""     # נימוק קצר למה התאים/לא התאים
    first_seen_year: int = 0   # השנה הראשונה שהאתר תועד ב-Wayback Machine (0=לא נמצא)
    domain_age_years: int = 0  # כמה שנים עברו מהסנאפשוט הראשון
    load_time_ms: int = 0      # זמן טעינה
    html_size_kb: float = 0.0  # גודל ה-HTML
    no_website: bool = False   # עסק בלי אתר אינטרנט
    social_url: str = ""       # קישור לדף פייסבוק/אינסטגרם/גוגל (כשאין אתר)
    # ----- תסריט שיחה אנושי -----
    script_intro: str = ""              # פתיחה — איך לפתוח את השיחה
    script_discovery: list = field(default_factory=list)  # 2-3 שאלות לגלות בעיות
    script_value_pitch: str = ""        # איך לחבר ערך לבעיה שלהם
    script_offer: str = ""              # ההצעה המעשית
    script_close: str = ""              # סגירה — מה לבקש
    script_objections: dict = field(default_factory=dict)  # {"אין זמן": "תשובה",...}
    script_dos_and_donts: list = field(default_factory=list)  # טיפים קצרים לשיחה


# ----------------------------------------------------------------------------
# שלב 1: חיפוש עסקים (מנגנונים מרובים, נופל אחד אחרי השני)
# ----------------------------------------------------------------------------

def _nominatim_bbox(city: str) -> Optional[tuple[float, float, float, float]]:
    """
    Use Nominatim to find a bounding box for the city in ISRAEL only.
    Returns (south, west, north, east) or None if not found in Israel.

    חשוב: אנחנו לא נופלים לחיפוש עולמי כי זה יכול להחזיר ערים עם
    שמות דומים במדינות אחרות (למשל 'הטילו' → Hatillo בפוארטו ריקו).
    """
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": city, "format": "json", "limit": 1, "countrycodes": "il"},
            headers=HEADERS,
            timeout=15,
        )
        if r.status_code != 200:
            return None
        data = r.json()
        if not data:
            # אם לא מצאנו את העיר בישראל - מחזירים None במקום לחפש בעולם.
            print(
                f"  ⚠️  Nominatim: '{city}' לא נמצאה בישראל. "
                f"בדקי איות (אפשר גם באנגלית).",
                file=sys.stderr,
            )
            return None
        bb = data[0].get("boundingbox")  # [south, north, west, east] as strings
        if not bb or len(bb) != 4:
            return None
        # אימות נוסף: וידוא שהתוצאה אכן בישראל (לפי country_code)
        country_code = (data[0].get("address", {}) or {}).get("country_code", "").lower()
        if country_code and country_code != "il":
            print(
                f"  ⚠️  Nominatim: '{city}' נמצאה במדינה {country_code}, לא בישראל. מתעלמים.",
                file=sys.stderr,
            )
            return None
        return (float(bb[0]), float(bb[2]), float(bb[1]), float(bb[3]))
    except Exception as e:
        print(f"  Nominatim failed: {e}", file=sys.stderr)
        return None


def search_businesses_overpass(
    city: str,
    business_type: str,
    limit: int = 50,
    *,
    only_without_website: bool = False,
) -> list[dict]:
    """
    Overpass API של OpenStreetMap - חינמי לחלוטין, ללא מפתח.
    כברירת מחדל מחזיר רק עסקים עם תג website.
    אם only_without_website=True - מחזיר רק עסקים בלי אתר (אבל עם טלפון).
    משתמש ב-Nominatim כדי לקבל bounding box של העיר (אמין יותר משימוש בשם בעברית).
    """
    # מילון מיפוי לתגי OSM
    type_map = {
        "מסעדה": '"amenity"~"restaurant|cafe|fast_food"',
        "restaurant": '"amenity"~"restaurant|cafe|fast_food"',
        "בית קפה": '"amenity"~"cafe|restaurant"',
        "cafe": '"amenity"="cafe"',
        "עורך דין": '"office"="lawyer"',
        "lawyer": '"office"="lawyer"',
        "רופא": '"amenity"="doctors"',
        "doctor": '"amenity"="doctors"',
        "רופא שיניים": '"amenity"="dentist"',
        "dentist": '"amenity"="dentist"',
        "מספרה": '"shop"="hairdresser"',
        "barber": '"shop"="hairdresser"',
        "חנות": '"shop"',
        "shop": '"shop"',
        "מלון": '"tourism"~"hotel|guest_house"',
        "hotel": '"tourism"~"hotel|guest_house"',
        "מוסך": '"shop"="car_repair"',
        "garage": '"shop"="car_repair"',
        "אופטיקה": '"shop"="optician"',
        "optician": '"shop"="optician"',
        "סטודיו": '"leisure"~"fitness_centre|sports_centre"',
        "gym": '"leisure"="fitness_centre"',
    }
    osm_filter = type_map.get(business_type.strip(), f'"name"~"{business_type}",i')

    # תגי OSM: או ["website"] (ברירת מחדל) או [!website][phone] (רק בלי אתר, חייב טלפון)
    if only_without_website:
        site_clause = '[!"website"][!"contact:website"]["phone"]'
    else:
        site_clause = '["website"]'

    # שלב א: נסה למצוא bounding box באמצעות Nominatim
    bbox = _nominatim_bbox(city)
    if bbox:
        s, w, n, e = bbox
        print(f"  Nominatim bbox for '{city}': {bbox} (only_without_website={only_without_website})", file=sys.stderr)
        query = f"""
        [out:json][timeout:30];
        (
          node[{osm_filter}]{site_clause}({s},{w},{n},{e});
          way[{osm_filter}]{site_clause}({s},{w},{n},{e});
        );
        out body {limit};
        """
    else:
        # גיבוי: שיטה ישנה בעזרת שם העיר
        print(f"  Nominatim לא הצליח, נופל לחיפוש לפי שם עיר", file=sys.stderr)
        query = f"""
        [out:json][timeout:30];
        (
          area["name"~"{city}",i]["boundary"="administrative"];
          area["name:he"~"{city}",i]["boundary"="administrative"];
          area["name:en"~"{city}",i]["boundary"="administrative"];
        )->.searchArea;
        (
          node[{osm_filter}]{site_clause}(area.searchArea);
          way[{osm_filter}]{site_clause}(area.searchArea);
        );
        out body {limit};
        """
    endpoints = [
        "https://overpass-api.de/api/interpreter",
        "https://overpass.kumi.systems/api/interpreter",
        "https://overpass.openstreetmap.ru/api/interpreter",
    ]
    for ep in endpoints:
        try:
            r = requests.post(ep, data={"data": query}, headers=HEADERS, timeout=45)
            if r.status_code == 200:
                data = r.json()
                results = []
                for el in data.get("elements", []):
                    tags = el.get("tags", {})
                    site = tags.get("website") or tags.get("contact:website") or ""
                    if only_without_website:
                        # אם יש בכל זאת אתר (יכול לקרות אם ה-API החזיר תגים נוספים) - דלג
                        if site:
                            continue
                        # חייב להיות טלפון כדי שיהיה אפשר לפנות
                        phone = tags.get("phone") or tags.get("contact:phone") or ""
                        if not phone:
                            continue
                        # אימות נוסף: חייב להיות טלפון ישראלי
                        if not _is_israeli_phone(phone):
                            continue
                        results.append({
                            "name": tags.get("name") or tags.get("name:he") or "Unknown",
                            "website": "",
                            "phone": phone,
                            "email": tags.get("email") or tags.get("contact:email") or "",
                            "address": ", ".join(filter(None, [
                                tags.get("addr:street"), tags.get("addr:housenumber"),
                                tags.get("addr:city")
                            ])),
                            "no_website": True,
                        })
                    else:
                        if not site:
                            continue
                        if not site.startswith(("http://", "https://")):
                            site = "http://" + site
                        results.append({
                            "name": tags.get("name") or tags.get("name:he") or "Unknown",
                            "website": site,
                            "phone": tags.get("phone") or tags.get("contact:phone") or "",
                            "email": tags.get("email") or tags.get("contact:email") or "",
                            "address": ", ".join(filter(None, [
                                tags.get("addr:street"), tags.get("addr:housenumber"),
                                tags.get("addr:city")
                            ])),
                        })
                tag = "without website" if only_without_website else "with website"
                print(f"  Overpass returned {len(results)} businesses {tag}", file=sys.stderr)
                if results:
                    return results[:limit]
        except Exception as e:
            print(f"  Overpass {ep} failed: {e}", file=sys.stderr)
    return []


def search_businesses_ddg(city: str, business_type: str, limit: int = 30) -> list[dict]:
    """גיבוי: חיפוש דרך DuckDuckGo (דורש: pip install duckduckgo-search)."""
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        print("  duckduckgo-search לא מותקן, מדלג על גיבוי זה", file=sys.stderr)
        return []
    query = f'{business_type} {city} site:co.il'
    results = []
    seen_domains = set()
    try:
        with DDGS() as ddgs:
            for r in ddgs.text(query, max_results=limit * 3):
                url = r.get("href", "")
                if not url:
                    continue
                domain = urlparse(url).netloc
                if domain in seen_domains:
                    continue
                # סינון: אנחנו לא רוצים לידים מאתרי קטלוג גדולים
                if any(x in domain for x in [
                    "facebook.com", "google.com", "yelp.com", "tripadvisor",
                    "wikipedia", "youtube.com", "linkedin.com", "instagram.com",
                    "easy.co.il", "b144.co.il", "144.co.il", "d.co.il", "zap.co.il",
                ]):
                    continue
                seen_domains.add(domain)
                results.append({
                    "name": r.get("title", domain),
                    "website": url,
                    "phone": "", "email": "", "address": "",
                })
                if len(results) >= limit:
                    break
    except Exception as e:
        print(f"  DDG search failed: {e}", file=sys.stderr)
    return results


_DOMAIN_BLACKLIST = {
    "facebook.com", "instagram.com", "youtube.com", "google.com",
    "yelp.com", "tripadvisor.com", "tripadvisor.co.il", "wikipedia.org",
    "linkedin.com", "twitter.com", "x.com", "tiktok.com",
    "easy.co.il", "b144.co.il", "144.co.il", "d.co.il", "zap.co.il",
    "rest.co.il", "mouse.co.il", "ynet.co.il", "walla.co.il",
    "haaretz.co.il", "calcalist.co.il", "themarker.com",
    "restaurants.co.il", "rol.co.il", "mako.co.il", "n12.co.il",
    "wcms.co.il", "lametayel.co.il",
}


def _domain_of(url: str) -> str:
    """Extract a normalized domain from a URL (no scheme, no www, no trailing path)."""
    if not url:
        return ""
    try:
        netloc = urlparse(url if "://" in url else f"http://{url}").netloc.lower()
    except Exception:
        return url.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return netloc


def _is_blacklisted_domain(url: str) -> bool:
    d = _domain_of(url)
    if not d:
        return True
    for bad in _DOMAIN_BLACKLIST:
        if d == bad or d.endswith("." + bad):
            return True
    return False


def search_businesses_ai(
    city: str,
    business_type: str,
    description: str = "",
    limit: int = 30,
    exclude_domains: Optional[set[str]] = None,
    *,
    only_without_website: bool = False,
) -> list[dict]:
    """
    חיפוש עסקים מקומיים בעזרת OpenAI (gpt-4o-mini-search-preview) עם גלישה באינטרנט.
    זה מנוע החיפוש החזק ביותר - הוא מחפש בגוגל, נכנס לאתרים, ומחזיר תוצאות אמיתיות.
    אם only_without_website=True - מחפש רק עסקים שאין להם אתר אינטרנט.
    """
    api_key = (os.environ.get("OPENAI_API_KEY") or "").strip()
    if not api_key:
        print("  AI search: אין OPENAI_API_KEY, מדלג", file=sys.stderr)
        return []
    try:
        from openai import OpenAI
    except ImportError:
        print("  AI search: חבילת openai לא מותקנת", file=sys.stderr)
        return []

    exclude_domains = exclude_domains or set()
    excluded_text = ""
    if exclude_domains:
        sample = list(exclude_domains)[:30]
        excluded_text = (
            "\n\nאתרים שכבר חיפשנו בעבר ויש להתעלם מהם (אל תחזירי אותם שוב):\n"
            + "\n".join(f"- {d}" for d in sample)
        )

    desc_text = ""
    if description.strip():
        desc_text = f"\n\nתיאור נוסף של מה שאני מחפשת:\n{description.strip()}"

    blacklist_text = ", ".join(sorted(_DOMAIN_BLACKLIST))

    if only_without_website:
        user_prompt = f"""אני מחפשת לידים — עסקים מקומיים בישראל בלבד מסוג "{business_type}" בעיר "{city}" (ישראל) שאין להם אתר אינטרנט עצמאי.{desc_text}{excluded_text}

חשוב מאוד — אסור להמציא נתונים:
1. ‏**ישראל בלבד** — אסור עסקים מחו"ל. אם יש עיר עם שם דומה במדינה אחרת (Hatillo בפוארטו ריקו, Tel Aviv במקום אחר וכו') — אסור! רק ישראל.
2. ‏**כל טלפון חייב להיות מאומת** — ראית אותו במקור אמיתי באינטרנט. אם לא ראית טלפון אמיתי — אל תכלילי את העסק.
3. ‏**חובה לכלול שדה `source_url`** — הקישור שבו ראית את הטלפון (למשל דף הפייסבוק/אינסטגרם/Google Maps של העסק). זה חיוני לאימות.
4. כל הטלפונים חייבים להיות ישראליים (קידומת 0XX או 972+). אסור +1, +44, או כל קידומת זרה.
5. מצאי לפחות {limit} עסקים שאין להם אתר עצמאי משלהם — רק דף בגוגל, פייסבוק, אינסטגרם, או רישום באתרי תיירות/דירוגים.
6. כתובת בעיר {city} ישראל בלבד.
7. אסור לחזור על אותו עסק.
8. דרכים למצוא: Google Maps של עסקים מקומיים בישראל, רשימות פייסבוק, אתרי דירוגים ישראליים (zap, restaurants.co.il, easy.co.il). תני עדיפות לעסקים קטנים-בינוניים.

החזירי JSON תקני בלבד (ללא טקסט נוסף, ללא markdown), במבנה הזה:
{{"businesses":[{{"name":"שם","phone":"05X-XXXXXXX","address":"רחוב, עיר","social":"facebook/instagram/google url אופציונלי","source_url":"https://www.facebook.com/business-page-where-you-saw-the-phone"}}, ...]}}

אל תכלילי שדה website. ‏**אסור להמציא טלפונים**. אם אין לך מקור אמיתי שמופיע בו טלפון — אל תכלילי את העסק. עדיף 5 לידים אמיתיים מאשר 20 ממציאים."""
    else:
        user_prompt = f"""אני מחפשת לידים — עסקים מקומיים בישראל בלבד מסוג "{business_type}" בעיר "{city}" (ישראל) שיש להם אתר אינטרנט פעיל.{desc_text}{excluded_text}

חשוב מאוד:
1. ‏**ישראל בלבד** — אסור עסקים מחו"ל. אם יש עיר עם שם דומה במדינה אחרת (Hatillo בפוארטו ריקו, Tel Aviv במקום אחר וכו') — אסור! רק ישראל.
2. עדיפות לאתרים בעברית או .co.il / .org.il. אם יש טלפון, חייב להיות ישראלי (0XX או 972+).
3. מצאי לפחות {limit} עסקים שונים. תחפשי בגוגל, תכנסי לרשימות עסקים, תאתרי אתרים אמיתיים.
4. כל עסק חייב URL של האתר העצמאי שלו (לא דף פייסבוק, לא דף ב-easy.co.il, לא ב-zap, לא ב-restaurants.co.il).
5. אסור לחזור על אותו עסק (לפי דומיין).
6. עדיפות לעסקים שאתרם נראה ישן/לא מקצועי — הם הלידים הכי שווים.
7. כל עסק חייב להיות באמת בעיר {city} ישראל, לא בעיר אחרת.

החזירי JSON תקני בלבד (ללא טקסט נוסף, ללא markdown), במבנה הזה:
{{"businesses":[{{"name":"שם","website":"https://...","phone":"","address":""}}, ...]}}

אם את לא בטוחה לגבי טלפון או כתובת — השאירי מחרוזת ריקה. אל תמציאי. דומיינים אסורים: {blacklist_text}."""

    try:
        client = OpenAI(api_key=api_key)
        # gpt-4o-mini-search-preview יודע לגלוש באינטרנט - הכלי הזה נועד בדיוק לשימוש כזה
        completion = client.chat.completions.create(
            model="gpt-4o-mini-search-preview",
            web_search_options={"search_context_size": "medium"},
            messages=[
                {"role": "system", "content": "את עוזרת חיפוש לידים מקומיים בישראל. את מחפשת באינטרנט, נכנסת לתוצאות, ומחזירה JSON תקני בלבד. לעולם אל תכתבי טקסט מחוץ ל-JSON."},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = completion.choices[0].message.content or ""
    except Exception as e:
        print(f"  AI search failed: {e}", file=sys.stderr)
        return []

    # מחלצת את ה-JSON מהתשובה (גם אם יש טקסט מסביב או markdown fences)
    import json as _json
    json_text = content
    # מנקה ```json ... ``` fences אם יש
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", json_text)
    if fence:
        json_text = fence.group(1)
    # מחפש את הבלוק הראשון שמתחיל ב-{
    brace = re.search(r"\{[\s\S]*\}", json_text)
    if brace:
        json_text = brace.group(0)

    try:
        data = _json.loads(json_text)
    except Exception as e:
        print(f"  AI returned non-JSON: {e}; preview={content[:200]!r}", file=sys.stderr)
        return []

    raw = data.get("businesses") or data.get("results") or []
    seen_keys: set[str] = set()
    out: list[dict] = []
    for b in raw:
        if not isinstance(b, dict):
            continue
        if only_without_website:
            phone = (b.get("phone") or "").strip()
            name = (b.get("name") or "").strip()
            social = (b.get("social") or "").strip()
            source_url = (b.get("source_url") or "").strip()
            if not phone or not name:
                continue
            # סינון: חובה טלפון ישראלי. עסקים מחו"ל נחתכים כאן.
            if not _is_israeli_phone(phone):
                print(f"  AI search: דילוג על '{name}' — טלפון לא ישראלי: {phone}", file=sys.stderr)
                continue
            # סינון: חובה שיהיה מקור אמין לטלפון (קישור URL).
            # זה הכלי העיקרי שלנו נגד AI שממציא מספרים.
            verifiable_url = source_url or social
            if not verifiable_url or not verifiable_url.startswith(("http://", "https://")):
                print(
                    f"  AI search: דילוג על '{name}' — אין מקור URL לאימות הטלפון",
                    file=sys.stderr,
                )
                continue
            # de-dup לפי שם+טלפון (אין לנו דומיין)
            phone_digits_dd = re.sub(r"\D", "", phone)
            key = f"{name.lower()}|{phone_digits_dd}"
            if key in seen_keys:
                continue
            seen_keys.add(key)
            out.append({
                "name": name,
                "website": "",
                "phone": phone,
                "email": (b.get("email") or "").strip(),
                "address": (b.get("address") or "").strip(),
                "social": social or source_url,
                "source_url": source_url or social,
                "no_website": True,
            })
            if len(out) >= limit:
                break
        else:
            site = (b.get("website") or b.get("url") or "").strip()
            if not site:
                continue
            if not site.startswith(("http://", "https://")):
                site = "https://" + site
            if _is_blacklisted_domain(site):
                continue
            d = _domain_of(site)
            if d in exclude_domains or d in seen_keys:
                continue
            # אם יש טלפון - חובה שיהיה ישראלי. אם הוא ריק - נסבול את זה.
            phone = (b.get("phone") or "").strip()
            if phone and not _is_israeli_phone(phone):
                print(f"  AI search: דילוג על '{site}' — טלפון לא ישראלי: {phone}", file=sys.stderr)
                continue
            seen_keys.add(d)
            out.append({
                "name": (b.get("name") or "").strip() or d,
                "website": site,
                "phone": phone,
                "email": (b.get("email") or "").strip(),
                "address": (b.get("address") or "").strip(),
            })
            if len(out) >= limit:
                break
    tag = "no-website" if only_without_website else "with-website"
    print(f"  AI search returned {len(out)} unique businesses ({tag})", file=sys.stderr)
    return out


def find_businesses(
    city: str,
    business_type: str,
    limit: int,
    description: str = "",
    exclude_domains: Optional[set[str]] = None,
    *,
    only_without_website: bool = False,
) -> list[dict]:
    """
    מאתר עסקים ממקורות שונים. מנסה קודם AI (החזק ביותר), ואז Overpass, ואז DuckDuckGo.
    משלב את כל התוצאות ומסיר כפילויות.
    אם only_without_website=True - מחזיר רק עסקים שאין להם אתר.
    """
    exclude_domains = set(exclude_domains or set())
    seen_keys: set[str] = set()  # דומיינים (כשיש אתר) או שם+טלפון (כשאין)
    combined: list[dict] = []

    def _key_of(r: dict) -> str:
        site = r.get("website", "")
        if site and not _is_blacklisted_domain(site):
            return _domain_of(site)
        # ללא אתר: dedup לפי שם + טלפון מנורמל
        name = (r.get("name") or "").strip().lower()
        phone = re.sub(r"\D", "", (r.get("phone") or ""))
        return f"{name}|{phone}"

    def _add_results(rs: list[dict], source: str):
        added = 0
        for r in rs:
            site = r.get("website", "")
            if only_without_website:
                # רק עסקים בלי אתר
                if site:
                    continue
                if not (r.get("phone") or "").strip():
                    continue
            else:
                # עסקים עם אתר בלבד
                if not site or _is_blacklisted_domain(site):
                    continue
                d = _domain_of(site)
                if d in exclude_domains:
                    continue
            key = _key_of(r)
            if key in seen_keys or key in exclude_domains:
                continue
            seen_keys.add(key)
            combined.append(r)
            added += 1
            if len(combined) >= limit:
                break
        print(f"  [{source}] תרם {added} עסקים חדשים (סה\"כ עד כה: {len(combined)})")

    mode_tag = "ללא אתר" if only_without_website else "עם אתר"
    print(f"מחפש '{business_type}' ב'{city}' ({mode_tag})...")
    if description:
        print(f"  תיאור: {description[:120]}")
    if exclude_domains:
        print(f"  מתעלם מ-{len(exclude_domains)} דומיינים שכבר חיפשנו")

    if only_without_website:
        # במצב "בלי אתר" — OSM קודם! זה נתונים אמיתיים מקהילת OpenStreetMap
        # (אנשים אמיתיים תרמו ידנית את הטלפונים). AI נוטה להמציא מספרים.
        print("  [1] מנסה OpenStreetMap Overpass (מקור אמין)...")
        osm_results = search_businesses_overpass(
            city, business_type, limit,
            only_without_website=True,
        )
        _add_results(osm_results, "OSM")
        if len(combined) >= limit:
            return combined[:limit]

        print("  [2] משלים עם AI (gpt-4o-mini-search-preview) - רק עם מקור מאומת...")
        ai_results = search_businesses_ai(
            city, business_type, description, limit,
            exclude_domains | seen_keys,
            only_without_website=True,
        )
        _add_results(ai_results, "AI")
    else:
        # במצב "עם אתר" — AI יעיל יותר כי יש URL לאמת
        print("  [1] מנסה AI (gpt-4o-mini-search-preview)...")
        ai_results = search_businesses_ai(
            city, business_type, description, limit,
            exclude_domains | seen_keys,
            only_without_website=False,
        )
        _add_results(ai_results, "AI")
        if len(combined) >= limit:
            return combined[:limit]

        print("  [2] משלים עם OpenStreetMap Overpass...")
        osm_results = search_businesses_overpass(
            city, business_type, limit,
            only_without_website=False,
        )
        _add_results(osm_results, "OSM")
        if len(combined) >= limit:
            return combined[:limit]

        print("  [3] משלים עם DuckDuckGo...")
        ddg_results = search_businesses_ddg(city, business_type, limit)
        _add_results(ddg_results, "DDG")

    if not combined:
        print("  ✗ לא נמצאו עסקים בשום מקור", file=sys.stderr)
    return combined[:limit]


# ----------------------------------------------------------------------------
# שלב 2: ניתוח אתר וחישוב ציון התיישנות
# ----------------------------------------------------------------------------

# כל בעיה מוסיפה נקודות לציון "ההתיישנות". יותר נקודות = יותר ישן/הזדמנות גדולה.
def get_first_snapshot_year(url: str) -> int:
    """
    מחזיר את השנה הראשונה שהאתר תועד ב-Wayback Machine (Internet Archive).
    זה proxy טוב למתי האתר עלה לאוויר. 0 אם לא נמצא.
    """
    if not url:
        return 0
    try:
        # CDX API - מחזיר את הסנאפשוט הראשון
        domain = _domain_of(url)
        if not domain:
            return 0
        cdx_url = f"http://web.archive.org/cdx/search/cdx?url={domain}&from=1996&to={datetime.now().year}&limit=1&output=json"
        r = requests.get(cdx_url, headers=HEADERS, timeout=12)
        if r.status_code != 200:
            return 0
        data = r.json()
        # פורמט: שורה ראשונה = headers, שורה שנייה = הסנאפשוט הראשון
        if len(data) < 2:
            return 0
        timestamp = data[1][1]  # YYYYMMDDhhmmss
        if len(timestamp) >= 4:
            return int(timestamp[:4])
    except Exception as e:
        print(f"  Wayback lookup failed for {url}: {e}", file=sys.stderr)
    return 0


def analyze_website(url: str) -> tuple[int, list[str], dict]:
    issues = []
    score = 0
    meta = {
        "final_url": url,
        "has_https": False,
        "is_mobile_friendly": False,
        "last_copyright": "",
        "cms": "",
        "load_time_ms": 0,
        "html_size_kb": 0,
        "first_seen_year": 0,
        "domain_age_years": 0,
    }

    try:
        t0 = time.time()
        r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        meta["load_time_ms"] = int((time.time() - t0) * 1000)
        meta["final_url"] = r.url
        meta["html_size_kb"] = round(len(r.content) / 1024, 1)
    except requests.exceptions.SSLError:
        score += 20; issues.append("בעיית SSL/HTTPS")
        try:
            r = requests.get(url.replace("https://", "http://"),
                             headers=HEADERS, timeout=REQUEST_TIMEOUT)
        except Exception as e:
            # גם כשנכשלים - ננסה לקבל גיל דומיין כדי שלפחות נראה משהו
            fy = get_first_snapshot_year(url)
            if fy:
                meta["first_seen_year"] = fy
                meta["domain_age_years"] = max(0, datetime.now().year - fy)
            return 0, [f"לא ניתן לגשת לאתר: {e}"], meta
    except Exception as e:
        fy = get_first_snapshot_year(url)
        if fy:
            meta["first_seen_year"] = fy
            meta["domain_age_years"] = max(0, datetime.now().year - fy)
        return 0, [f"לא ניתן לגשת לאתר: {e}"], meta

    html = r.text
    soup = BeautifulSoup(html, "html.parser")

    # --- HTTPS ---
    meta["has_https"] = meta["final_url"].startswith("https://")
    if not meta["has_https"]:
        score += 15; issues.append("ללא HTTPS (אתר לא מאובטח)")

    # --- Viewport / מובייל ---
    viewport = soup.find("meta", attrs={"name": re.compile("viewport", re.I)})
    meta["is_mobile_friendly"] = bool(viewport)
    if not viewport:
        score += 25; issues.append("לא מותאם למובייל (אין viewport)")

    # --- תגיות HTML4 מיושנות ---
    legacy_tags = ["font", "center", "marquee", "blink", "frameset", "frame", "applet"]
    for tag in legacy_tags:
        if soup.find(tag):
            score += 8; issues.append(f"שימוש בתגית מיושנת <{tag}>")

    # --- Layout מבוסס טבלאות ---
    tables = soup.find_all("table")
    if len(tables) >= 3:
        nested = sum(1 for t in tables if t.find("table"))
        if nested >= 1 or len(tables) >= 5:
            score += 20; issues.append("עיצוב מבוסס טבלאות (ישן מאוד)")

    # --- Flash ---
    if soup.find("embed", src=re.compile(r"\.swf", re.I)) or \
       soup.find("object", attrs={"type": "application/x-shockwave-flash"}):
        score += 30; issues.append("שימוש ב-Flash (טכנולוגיה מתה)")

    # --- jQuery ישן ---
    for s in soup.find_all("script", src=True):
        src = s["src"]
        m = re.search(r"jquery[/-]?(\d+)\.(\d+)", src, re.I)
        if m:
            major = int(m.group(1))
            if major < 3:
                score += 10; issues.append(f"jQuery ישן ({m.group(0)})")
            break

    # --- Copyright ישן ---
    text = soup.get_text(" ", strip=True)
    years = re.findall(r"©\s*(\d{4})|copyright\s*(\d{4})|כל הזכויות שמורות\s*(\d{4})",
                       text, re.I)
    flat = [int(y) for tup in years for y in tup if y]
    if flat:
        latest = max(flat)
        meta["last_copyright"] = str(latest)
        current_year = datetime.now().year
        if latest < current_year - 4:
            score += 15; issues.append(f"Copyright ישן ({latest})")
        elif latest < current_year - 2:
            score += 8; issues.append(f"Copyright לא עדכני ({latest})")

    # --- WordPress ישן (אם זוהה) ---
    if 'wp-content' in html or 'wp-includes' in html:
        meta["cms"] = "WordPress"
        m = re.search(r'meta name="generator" content="WordPress (\d+)\.(\d+)', html)
        if m and int(m.group(1)) < 6:
            score += 10; issues.append(f"WordPress ישן ({m.group(1)}.{m.group(2)})")
    elif 'wix.com' in html:
        meta["cms"] = "Wix"
    elif 'shopify' in html.lower():
        meta["cms"] = "Shopify"
    elif 'squarespace' in html.lower():
        meta["cms"] = "Squarespace"

    # --- "best viewed in IE", hit counter, וסימנים אגדיים ---
    bad_phrases = [
        "best viewed in internet explorer", "optimized for ie",
        "hit counter", "מבקרים", "you are visitor",
        "under construction", "באתר בבנייה",
    ]
    lower_text = text.lower()
    for phrase in bad_phrases:
        if phrase in lower_text:
            score += 15; issues.append(f"טקסט מיושן: '{phrase}'"); break

    # --- ללא Open Graph / Schema.org ---
    if not soup.find("meta", attrs={"property": re.compile("^og:", re.I)}):
        score += 8; issues.append("ללא Open Graph (חסר SEO/שיתוף)")
    if not soup.find("script", attrs={"type": "application/ld+json"}):
        score += 5; issues.append("ללא Schema.org (חסר SEO מתקדם)")

    # --- ללא favicon ---
    if not soup.find("link", rel=re.compile("icon", re.I)):
        score += 4; issues.append("ללא favicon")

    # --- HTML מאוד קטן (סטטי ישן) ---
    if meta["html_size_kb"] < 5:
        score += 5; issues.append(f"HTML קטן מאוד ({meta['html_size_kb']}KB)")

    # --- טעינה איטית ---
    if meta["load_time_ms"] > 5000:
        score += 10; issues.append(f"טעינה איטית ({meta['load_time_ms']}ms)")

    # --- פונטים מיושנים ב-style inline ---
    if re.search(r'font-family\s*:\s*["\']?(comic sans|times new roman)',
                 html, re.I):
        score += 5; issues.append("פונט מיושן (Comic Sans / Times New Roman)")

    # --- שנת יצירת האתר (Wayback Machine) ---
    first_year = get_first_snapshot_year(meta["final_url"])
    if first_year:
        meta["first_seen_year"] = first_year
        age = max(0, datetime.now().year - first_year)
        meta["domain_age_years"] = age
        # אם האתר ישן וגם לא עבר עדכון משמעותי (אין https / mobile / og) - הזדמנות זהב
        if age >= 15:
            score += 15; issues.append(f"אתר ישן מאוד — קיים מ-{first_year} ({age} שנים)")
        elif age >= 10:
            score += 10; issues.append(f"אתר ישן — קיים מ-{first_year} ({age} שנים)")
        elif age >= 7:
            score += 5; issues.append(f"אתר לא חדש — קיים מ-{first_year} ({age} שנים)")

    # טקסט מעובד לניתוח AI (חיסכון בטוקנים)
    try:
        meta["page_title"] = (soup.title.string or "").strip() if soup.title else ""
        digest_src = soup.get_text("\n", strip=True)
        digest_src = re.sub(r"\n{3,}", "\n\n", digest_src)
        digest_src = re.sub(r"[ \t]{2,}", " ", digest_src)
        max_ai = 48_000
        if len(digest_src) > max_ai:
            digest_src = digest_src[:max_ai] + "\n…[נחתך לניתוח]"
        meta["html_for_ai"] = digest_src
    except Exception:
        meta["page_title"] = ""
        meta["html_for_ai"] = re.sub(r"\s+", " ", text)[:20_000] if text else ""

    return score, issues, meta


def score_to_grade(score: int) -> str:
    if score >= 80: return "F"   # ישן מאוד - הזדמנות זהב
    if score >= 60: return "D"   # ישן
    if score >= 40: return "C"   # בינוני
    if score >= 20: return "B"   # טוב
    return "A"                    # מעולה - לא כדאי לפנות


# ----------------------------------------------------------------------------
# ניתוח AI + הכנה לשיחה (אנושי בלבד — ללא שליחת הודעות אוטומטית)
# ----------------------------------------------------------------------------

_AI_JSON_INSTRUCTIONS = """את עוזרת מכירות ישראלית מנוסה. את עוזרת לבעלת עסק קטן (פרילנסרית) למכור שירותי בניית/שדרוג אתרים לעסקים.
החזירי JSON תקף בלבד, בלי טקסט לפני או אחרי. כל הטקסטים בעברית מדוברת, חמה ויומיומית — לא שיווק מנופח, לא מילים גבוהות, כמו שאחת חברה הייתה מסבירה לחברה. דברי בגוף ראשון יחיד (אני).

מפתחות:
{
  "summary": "משפט אחד עד שניים בעברית — איך האתר מרגיש",
  "ux_issues": ["בעיה UX קצרה"],
  "conversion_issues": ["בעיית המרה"],
  "trust_issues": ["בעיית אמון"],
  "main_problems": ["3–5 בעיות עסקיות קצרות"],
  "best_talking_point": "משפט טבעי לשיחה — לא שיווק מוגזם",
  "suggested_angle": "משפט אחד איך להתקרב",
  "opportunity_score": 0,
  "close_probability": 0,
  "strongest_problem": "הבעיה העסקית החזקה ביותר באתר — משפט אחד",
  "business_impact": "למה זה פוגע בעסק (פניות, אמון, זמן) — עד שני משפטים",
  "opening_line": "משפט פתיחה לשיחה — יכול להיות זהה ל-best_talking_point",
  "if_not_interested": "מה להגיד אם אומרים לא מעניין / אין זמן — משפט אחד קצר",
  "what_to_offer": "מה להציע במפורש (למשל: דף נחיתה, אתר חדש, שדרוג נייד) — משפט אחד",
  "next_action": "אחד מהבאים בלבד: call או whatsapp או skip",
  "match_score": 0,
  "match_reason": "אם נתון לי תיאור של מי מחפשים — כמה זה מתאים (0=לא מתאים, 100=התאמה מושלמת); משפט קצר למה. אם לא נתון תיאור — החזר 0 ומחרוזת ריקה.",

  "script_intro": "איך לפתוח את השיחה. 2-3 משפטים, ידידותי, לא מוכר מהדקה הראשונה. תני בדיוק את המילים שצריך להגיד ('היי, מדבר... ראיתי שיש לכם אתר ל...'). חשוב: להציג את עצמי, להגיד למה אני מתקשרת, ולשאול אם זה זמן טוב לדקה.",
  "script_discovery": ["2-3 שאלות שאני שואלת אותם כדי לגלות בעיות. שאלות אמיתיות שיגרמו להם להבין שיש להם בעיה. למשל: 'כמה פניות אתם מקבלים מהאתר בשבוע?' או 'מתי בפעם האחרונה עדכנתם את האתר?'. כל פריט במערך הוא שאלה אחת מנוסחת מילה במילה."],
  "script_value_pitch": "ברגע שגיליתי בעיה — איך אני מציגה את הערך שלי. 2-3 משפטים שמחברים את הבעיה שלהם לפתרון שלי. דבר על תוצאה עסקית (יותר פניות, יותר לקוחות, פחות בעיות), לא על טכנולוגיה.",
  "script_offer": "ההצעה הקונקרטית — 'אני יכולה לבנות לך X תוך Y שבועות ב-Z שקלים'. או 'בואי נקבע פגישה של 15 דקות בה אני אראה לך איך זה ייראה'. מילים מדויקות.",
  "script_close": "הקריאה לפעולה: מה אני מבקשת בסוף השיחה. למשל: 'אז מה דעתך שניפגש ביום שני בבוקר?' או 'אשלח לך הצעת מחיר במייל - מה האימייל הכי טוב?'. מנוסח כשאלה סגורה.",
  "script_objections": {
    "אין לי זמן": "תשובה אנושית של 1-2 משפטים — מבינה אותם, מציעה לחזור בזמן יותר נח",
    "יש לי כבר מישהו": "תשובה — לא דוחפת, מציעה ערך שונה (חוות דעת שניה, ייעוץ חינם)",
    "זה יקר מדי": "תשובה — שואלת מה הם משלמים היום או מציעה אפשרות זולה יותר",
    "לא מעוניין": "תשובה — שואלת מה הסיבה, מציעה לחזור בעתיד"
  },
  "script_dos_and_donts": ["3-5 טיפים קצרים — למשל: 'אל תקראי מהדף, תהיי טבעית', 'תני להם לדבר 70% מהזמן', 'אם הם שותקים — את שותקת', 'אל תפחדי להגיד את המחיר'"]
}

חוקי ניסוח קריטיים לתסריט השיחה:
1. ‏**עברית מדוברת**, לא ספרותית. כמו שמדברים, לא כמו שכותבים.
2. ‏**אסור** מילים כמו 'פתרון מקיף', 'טכנולוגיה מתקדמת', 'ערך מוסף'.
3. ‏**כן**: 'בואי נדבר על זה', 'אני אכין לך משהו', 'יש דרך פשוטה'.
4. שיחה כמו עם חבר, לא כמו דוקומנט שיווקי.
5. בכל פריט תני **מילים בדיוק כמו שצריך להגיד** ('היי X, מדבר/ת Y מ...'). אל תכתבי הוראות כמו 'תציגי את עצמך' — תני את הטקסט עצמו.
6. הכל אישי לעסק הזה — תשתמשי בשם העסק, סוג העסק, והבעיות הספציפיות שגילית.

opportunity_score, close_probability, match_score — מספרים שלמים 0–100. next_action חייב להיות בדיוק call, whatsapp או skip.
כל הרשימות עד 5 פריטים. עברית בלבד."""


def _empty_ai_result(reason: str = "") -> dict:
    return {
        "summary": "",
        "ux_issues": [],
        "conversion_issues": [],
        "trust_issues": [],
        "main_problems": [],
        "best_talking_point": "",
        "suggested_angle": "",
        "opportunity_score": 0,
        "close_probability": 0,
        "strongest_problem": "",
        "business_impact": "",
        "opening_line": "",
        "if_not_interested": "",
        "what_to_offer": "",
        "next_action": "",
        "match_score": 0,
        "match_reason": "",
        "script_intro": "",
        "script_discovery": [],
        "script_value_pitch": "",
        "script_offer": "",
        "script_close": "",
        "script_objections": {},
        "script_dos_and_donts": [],
        "_reason": reason,
    }


def _clamp_int(v, lo: int = 0, hi: int = 100) -> int:
    try:
        x = int(float(v))
        return max(lo, min(hi, x))
    except (TypeError, ValueError):
        return 0


def _normalize_next_action(v: str) -> str:
    s = (v or "").strip().lower()
    if s in ("call", "whatsapp", "skip"):
        return s
    if "wa" in s or "ווצ" in s:
        return "whatsapp"
    if "דלג" in s or "skip" in s:
        return "skip"
    return ""


def analyze_with_ai(html: str, url: str, description: str = "") -> dict:
    """
    מנתח את תוכן האתר (טקסט שחולץ מה-HTML) ומחזיר תובנות לשיחה אנושית.
    דורש OPENAI_API_KEY בסביבה. לא שולח הודעות ולא מבצע אאוטריץ'.
    אם נתון `description` — ה-AI גם יחזיר match_score (0-100) ו-match_reason לפי ההתאמה לתיאור.
    """
    key = (os.environ.get("OPENAI_API_KEY") or "").strip()
    if not key:
        d = _empty_ai_result("no_api_key")
        d["_reason"] = "חסר OPENAI_API_KEY — ניתוח AI דולג"
        return d
    digest = (html or "").strip()
    if len(digest) < 80:
        d = _empty_ai_result("too_short")
        d["_reason"] = "מעט מדי תוכן לניתוח AI"
        return d
    try:
        from openai import OpenAI
    except ImportError:
        d = _empty_ai_result("no_openai_pkg")
        d["_reason"] = "חבילת openai לא מותקנת (pip install openai)"
        return d

    client = OpenAI(api_key=key)
    model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
    desc_part = ""
    if (description or "").strip():
        desc_part = (
            f"\n\nתיאור הליד שהמשתמש מחפש (שתשתמשי בו לקביעת התאמה):\n"
            f"---\n{description.strip()[:1500]}\n---\n"
            f"החזירי גם match_score (0-100) - כמה הליד הזה מתאים לתיאור, "
            f"ו-match_reason - משפט קצר למה."
        )
    user_blob = (
        f"כתובת האתר: {url}\n\n"
        f"תוכן טקסטואלי שחולץ מהדף (ייתכן חתוך):\n---\n{digest[:42_000]}\n---"
        f"{desc_part}"
    )
    try:
        resp = client.chat.completions.create(
            model=model,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": _AI_JSON_INSTRUCTIONS},
                {"role": "user", "content": user_blob},
            ],
            temperature=0.4,
            max_tokens=2_000,
        )
        raw = (resp.choices[0].message.content or "").strip()
        data = json.loads(raw)
    except Exception as e:
        d = _empty_ai_result("api_error")
        d["_reason"] = f"שגיאת API: {e}"
        return d

    # פענוח מילון "התנגדויות" - תמיד dict של מחרוזת→מחרוזת
    raw_obj = data.get("script_objections") or {}
    objections: dict[str, str] = {}
    if isinstance(raw_obj, dict):
        for k, v in raw_obj.items():
            ks = str(k).strip()
            vs = str(v).strip()
            if ks and vs:
                objections[ks] = vs
    elif isinstance(raw_obj, list):
        # בעבר היה לפעמים מערך - נתמוך גם בפורמט הישן
        for item in raw_obj:
            if isinstance(item, dict):
                k = str(item.get("objection") or item.get("q") or "").strip()
                v = str(item.get("response") or item.get("a") or "").strip()
                if k and v:
                    objections[k] = v

    out = {
        "summary": str(data.get("summary") or "").strip(),
        "ux_issues": [str(x).strip() for x in (data.get("ux_issues") or []) if str(x).strip()][:5],
        "conversion_issues": [str(x).strip() for x in (data.get("conversion_issues") or []) if str(x).strip()][:5],
        "trust_issues": [str(x).strip() for x in (data.get("trust_issues") or []) if str(x).strip()][:5],
        "main_problems": [str(x).strip() for x in (data.get("main_problems") or []) if str(x).strip()][:5],
        "best_talking_point": str(data.get("best_talking_point") or "").strip(),
        "suggested_angle": str(data.get("suggested_angle") or "").strip(),
        "opportunity_score": _clamp_int(data.get("opportunity_score")),
        "close_probability": _clamp_int(data.get("close_probability")),
        "strongest_problem": str(data.get("strongest_problem") or "").strip(),
        "business_impact": str(data.get("business_impact") or "").strip(),
        "opening_line": str(data.get("opening_line") or "").strip(),
        "if_not_interested": str(data.get("if_not_interested") or "").strip(),
        "what_to_offer": str(data.get("what_to_offer") or "").strip(),
        "next_action": _normalize_next_action(str(data.get("next_action") or "")),
        "match_score": _clamp_int(data.get("match_score")) if data.get("match_score") is not None else 0,
        "match_reason": str(data.get("match_reason") or "").strip(),
        "script_intro": str(data.get("script_intro") or "").strip(),
        "script_discovery": [str(x).strip() for x in (data.get("script_discovery") or []) if str(x).strip()][:5],
        "script_value_pitch": str(data.get("script_value_pitch") or "").strip(),
        "script_offer": str(data.get("script_offer") or "").strip(),
        "script_close": str(data.get("script_close") or "").strip(),
        "script_objections": objections,
        "script_dos_and_donts": [str(x).strip() for x in (data.get("script_dos_and_donts") or []) if str(x).strip()][:6],
    }
    if len(out["main_problems"]) > 5:
        out["main_problems"] = out["main_problems"][:5]
    return out


def apply_ai_dict_to_lead(lead: Lead, d: dict) -> None:
    lead.ai_summary = d.get("summary") or ""
    lead.ux_issues = list(d.get("ux_issues") or [])
    lead.conversion_issues = list(d.get("conversion_issues") or [])
    lead.trust_issues = list(d.get("trust_issues") or [])
    mp = list(d.get("main_problems") or [])
    lead.main_problems = mp[:5]
    lead.best_talking_point = d.get("best_talking_point") or ""
    lead.suggested_angle = d.get("suggested_angle") or ""
    if d.get("opportunity_score") is not None:
        lead.opportunity_score = _clamp_int(d.get("opportunity_score"))
    if d.get("close_probability") is not None:
        lead.close_probability = _clamp_int(d.get("close_probability"))
    lead.strongest_problem = d.get("strongest_problem") or ""
    lead.business_impact = d.get("business_impact") or ""
    lead.opening_line = d.get("opening_line") or ""
    lead.if_not_interested = d.get("if_not_interested") or ""
    lead.what_to_offer = d.get("what_to_offer") or ""
    na = _normalize_next_action(d.get("next_action") or "")
    if na:
        lead.next_action = na
    if d.get("match_score") is not None:
        lead.match_score = _clamp_int(d.get("match_score"))
    if d.get("match_reason"):
        lead.match_reason = str(d.get("match_reason")).strip()
    # תסריט שיחה אנושי
    lead.script_intro = d.get("script_intro") or ""
    lead.script_discovery = list(d.get("script_discovery") or [])
    lead.script_value_pitch = d.get("script_value_pitch") or ""
    lead.script_offer = d.get("script_offer") or ""
    lead.script_close = d.get("script_close") or ""
    obj = d.get("script_objections")
    lead.script_objections = dict(obj) if isinstance(obj, dict) else {}
    lead.script_dos_and_donts = list(d.get("script_dos_and_donts") or [])
    r = d.get("_reason")
    if not r:
        lead.ai_notes = ""
    elif r == "no_api_key":
        lead.ai_notes = "ניתוח AI לא רץ — הגדרי OPENAI_API_KEY"
    elif r == "too_short":
        lead.ai_notes = "לא ניתן היה לנתח — מעט מדי טקסט מהאתר"
    elif r == "no_openai_pkg":
        lead.ai_notes = "התקיני: pip install openai"
    else:
        lead.ai_notes = str(r)


def compute_priority_level(lead: Lead) -> str:
    """
    high: אתר בעייתי + יש טלפון/וואטסאפ + בעיות ברורות
    medium: אתר בינוני או חסר קשר
    low: אתר טוב יחסית או קשה להגיע
    """
    has_contact = bool(
        re.sub(r"\D", "", (lead.phone or "").strip())
        or (lead.whatsapp or "").strip()
    )
    bad_site = lead.grade in ("F", "D") or lead.score >= 55
    good_site = lead.grade in ("A", "B") or (lead.score < 28 and not lead.error)
    n_prob = len(lead.main_problems or [])
    n_heur = len(lead.issues or [])
    clear_problems = n_prob >= 2 or (bad_site and n_heur >= 3) or (n_prob >= 1 and bad_site)

    if good_site:
        return "low"
    if bad_site and has_contact and clear_problems:
        return "high"
    if not has_contact:
        return "low" if good_site else "medium"
    if bad_site and not clear_problems:
        return "medium"
    if 40 <= lead.score < 55:
        return "medium"
    return "medium"


def enrich_lead_for_crm(lead: Lead) -> None:
    """משלים שדות קרב/סקור אם חסרים (אחרי AI או בלי)."""
    if not lead.opening_line and (lead.best_talking_point or "").strip():
        lead.opening_line = lead.best_talking_point.strip()
    if not lead.strongest_problem:
        if lead.main_problems:
            lead.strongest_problem = str(lead.main_problems[0])
        elif lead.issues:
            lead.strongest_problem = str(lead.issues[0])
        else:
            lead.strongest_problem = "האתר לא מנצל טוב את התנועה מהנייד והמסר לא מספיק ברור"
    if not lead.business_impact:
        lead.business_impact = (
            "גולשים נוטשים מהר כשהחוויה מבלבלת או איטית — זה עלול לצמצם פניות "
            "ותדמית מקצועית מול מתחרים."
        )
    if not lead.if_not_interested:
        lead.if_not_interested = (
            "בסדר גמור. אם תרצו בעתיד משוב קצר בלי התחייבות — אשמח לעזור."
        )
    if not lead.what_to_offer:
        lead.what_to_offer = (
            "אתר מודרני עם דגש נייד, כפתורי קשר בולטים, ומסר שמוביל לפעולה."
        )
    has_p = bool(re.sub(r"\D", "", (lead.phone or "").strip()))
    has_w = bool((lead.whatsapp or "").strip())
    if not (lead.next_action or "").strip():
        if lead.priority_level == "low" or lead.grade in ("A", "B"):
            lead.next_action = "skip"
        elif has_w:
            lead.next_action = "whatsapp"
        elif has_p:
            lead.next_action = "call"
        else:
            lead.next_action = "skip"
    if lead.opportunity_score <= 0:
        o = min(100, int(lead.score * 0.62 + (22 if has_p else 0) + (14 if has_w else 0)))
        if lead.grade == "A":
            o -= 38
        elif lead.grade == "B":
            o -= 14
        lead.opportunity_score = max(5, min(100, o))
    if lead.close_probability <= 0:
        if lead.priority_level == "high":
            lead.close_probability = 50
        elif lead.priority_level == "medium":
            lead.close_probability = 34
        else:
            lead.close_probability = 18


def generate_call_prep(lead: Lead) -> str:
    """טקסט קצר להדפסה / הדבקה לפני שיחה — לא נשלח אוטומטית."""
    probs = list(lead.main_problems or [])[:3]
    if len(probs) < 2:
        for x in (lead.issues or []):
            if x not in probs:
                probs.append(x)
            if len(probs) >= 3:
                break
    bullets = "\n".join(f"- {p}" for p in probs[:3]) or "- (אין נקודות — בדקי את האתר ידנית)"
    tp = (lead.best_talking_point or "").strip() or "אפשר לפתוח בקצרה: ראיתי את האתר ויש נקודה אחת שכנראה מורידה לכם פניות מהנייד — רוצים שאפרט בשתי דקות?"
    angle = (lead.suggested_angle or "").strip() or "התמקדי בערך עסקי קונקרטי (זמן, פניות, ביטחון) — בלי לחץ."
    name = (lead.business_name or "העסק").strip()
    return (
        f"עסק: {name}\n\n"
        f"מה לא עובד:\n{bullets}\n\n"
        f"מה לומר (משפט אחד):\n{tp}\n\n"
        f"איך להתקרב:\n{angle}\n\n"
        f"מטרת השיחה:\n"
        f"שיאמרו: «כן, שלחי לי מה שאת מציעה»\n"
    )


def capture_homepage_screenshot(url: str, dest_path: Path) -> bool:
    """אופציונלי: דורש playwright + chromium (playwright install chromium)."""
    href = _url_for_href(url)
    if not href:
        return False
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return False
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            try:
                page = browser.new_page(viewport={"width": 1280, "height": 800})
                page.goto(href, timeout=25_000, wait_until="domcontentloaded")
                page.wait_for_timeout(800)
                page.screenshot(path=str(dest_path), full_page=False)
            finally:
                browser.close()
    except Exception:
        return False
    return True


def lead_to_dashboard_dict(lead: Lead, html_output_path: str) -> dict:
    """מבנה לדשבורד HTML (JSON) כולל נתיב יחסי לצילום."""
    shot = ""
    if lead.screenshot_path:
        try:
            shot = str(Path(lead.screenshot_path).relative_to(Path(html_output_path).parent))
        except ValueError:
            shot = Path(lead.screenshot_path).name
    return {
        "business_name": lead.business_name,
        "website": lead.website,
        "final_url": lead.final_url or lead.website,
        "score": lead.score,
        "grade": lead.grade,
        "priority_level": lead.priority_level,
        "phone": lead.phone,
        "whatsapp": lead.whatsapp,
        "email": lead.email,
        "error": lead.error,
        "ai_summary": lead.ai_summary,
        "main_problems": lead.main_problems,
        "ux_issues": lead.ux_issues,
        "trust_issues": lead.trust_issues,
        "conversion_issues": lead.conversion_issues,
        "best_talking_point": lead.best_talking_point,
        "suggested_angle": lead.suggested_angle,
        "ai_notes": lead.ai_notes,
        "call_prep": generate_call_prep(lead),
        "screenshot_rel": shot,
        "search_city": lead.search_city,
        "search_business_type": lead.search_business_type,
        "opportunity_score": lead.opportunity_score,
        "close_probability": lead.close_probability,
        "strongest_problem": lead.strongest_problem,
        "business_impact": lead.business_impact,
        "opening_line": lead.opening_line,
        "if_not_interested": lead.if_not_interested,
        "what_to_offer": lead.what_to_offer,
        "next_action": lead.next_action,
    }


def lead_to_supabase_payload(lead: Lead, *, user_id: str, job_id: str) -> dict:
    """מיפוי לשורה ב-Supabase (עמודות טבלת leads)."""
    now_iso = datetime.utcnow().isoformat() + "Z"
    final_norm = _normalize_url_for_storage(lead.final_url or lead.website)
    site_key = lead.site_key or _site_key_from_url(final_norm)
    phone_norm = _normalize_il_phone(lead.phone or "")
    wa_norm = re.sub(r"\D", "", (lead.whatsapp or ""))
    return {
        "user_id": user_id,
        "search_job_id": job_id,
        "business_name": lead.business_name or None,
        "website": _normalize_url_for_storage(lead.website) or None,
        "final_url": final_norm or None,
        "score": lead.score,
        "grade": lead.grade or None,
        "issues": lead.issues or [],
        "email": lead.email or None,
        "phone": phone_norm or None,
        "whatsapp": wa_norm or None,
        "facebook": lead.facebook or None,
        "instagram": lead.instagram or None,
        "address": lead.address or None,
        "last_copyright": lead.last_copyright or None,
        "has_https": lead.has_https,
        "is_mobile_friendly": lead.is_mobile_friendly,
        "cms": lead.cms or None,
        "lead_error": lead.error or None,
        "error_reason": lead.error or None,
        "ai_summary": lead.ai_summary or None,
        "main_problems": lead.main_problems or [],
        "ux_issues": lead.ux_issues or [],
        "trust_issues": lead.trust_issues or [],
        "conversion_issues": lead.conversion_issues or [],
        "best_talking_point": lead.best_talking_point or None,
        "suggested_angle": lead.suggested_angle or None,
        "priority_level": lead.priority_level or None,
        "ai_notes": lead.ai_notes or None,
        "screenshot_path": lead.screenshot_path or None,
        "search_city": lead.search_city or None,
        "search_business_type": lead.search_business_type or None,
        "opportunity_score": lead.opportunity_score,
        "close_probability": lead.close_probability,
        "strongest_problem": lead.strongest_problem or None,
        "business_impact": lead.business_impact or None,
        "opening_line": lead.opening_line or None,
        "if_not_interested": lead.if_not_interested or None,
        "what_to_offer": lead.what_to_offer or None,
        "next_action": lead.next_action or None,
        "call_prep": generate_call_prep(lead),
        "status": "new",
        "site_key": site_key or None,
        "last_analyzed_at": now_iso,
        "match_score": lead.match_score or 0,
        "match_reason": lead.match_reason or None,
        "first_seen_year": lead.first_seen_year or 0,
        "domain_age_years": lead.domain_age_years or 0,
        "load_time_ms": lead.load_time_ms or 0,
        "html_size_kb": float(lead.html_size_kb or 0),
        "no_website": lead.no_website,
        "social_url": lead.social_url or None,
        "script_intro": lead.script_intro or None,
        "script_discovery": lead.script_discovery or [],
        "script_value_pitch": lead.script_value_pitch or None,
        "script_offer": lead.script_offer or None,
        "script_close": lead.script_close or None,
        "script_objections": lead.script_objections or {},
        "script_dos_and_donts": lead.script_dos_and_donts or [],
    }


# ----------------------------------------------------------------------------
# שלב 3: חילוץ פרטי קשר
# ----------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[\w._%+-]+@[\w.-]+\.[A-Za-z]{2,6}")
IL_PHONE_RE = re.compile(r"(?:\+972[-\s]?|0)([23489]|5[023456789])[-\s]?\d{3}[-\s]?\d{4}")
WA_RE = re.compile(r"(?:wa\.me/|api\.whatsapp\.com/send\?phone=)(\+?\d+)")
FB_RE = re.compile(r"facebook\.com/([A-Za-z0-9.\-_]+)")
IG_RE = re.compile(r"instagram\.com/([A-Za-z0-9._]+)")

GENERIC_EMAILS = {"example", "no-reply", "noreply", "test", "your", "admin@admin"}

def extract_contacts(url: str, html: str = None) -> dict:
    """ אם html סופק - מנתח אותו. אחרת מוריד את הדף ובודק /contact /צור-קשר """
    contacts = {"email": "", "phone": "", "whatsapp": "", "facebook": "", "instagram": ""}
    sources = [(url, html)] if html else []
    if not html:
        try:
            r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            sources.append((url, r.text))
        except Exception:
            return contacts

    # דף "צור קשר"
    try:
        soup = BeautifulSoup(sources[0][1], "html.parser")
        for a in soup.find_all("a", href=True):
            h = a["href"].lower()
            text = (a.get_text() or "").strip()
            if any(k in h for k in ["contact", "צור-קשר", "tzor"]) or \
               any(k in text for k in ["צור קשר", "Contact", "יצירת קשר"]):
                full = urljoin(url, a["href"])
                if full not in [s[0] for s in sources]:
                    try:
                        rr = requests.get(full, headers=HEADERS, timeout=REQUEST_TIMEOUT)
                        sources.append((full, rr.text))
                        break
                    except Exception:
                        pass
    except Exception:
        pass

    combined = " ".join(s[1] for s in sources)

    # אימייל
    for m in EMAIL_RE.findall(combined):
        if not any(g in m.lower() for g in GENERIC_EMAILS) and \
           not m.lower().endswith((".png", ".jpg", ".gif", ".svg")):
            contacts["email"] = m; break

    # טלפון
    m = IL_PHONE_RE.search(combined)
    if m:
        contacts["phone"] = m.group(0)

    # WhatsApp
    m = WA_RE.search(combined)
    if m:
        contacts["whatsapp"] = m.group(1)

    # פייסבוק / אינסטגרם
    m = FB_RE.search(combined)
    if m and m.group(1) not in ("sharer", "tr", "plugins", "dialog"):
        contacts["facebook"] = m.group(1)
    m = IG_RE.search(combined)
    if m and m.group(1) not in ("p", "explore", "reel"):
        contacts["instagram"] = m.group(1)

    return contacts


# ----------------------------------------------------------------------------
# שלב 4: ייצוא לאקסל
# ----------------------------------------------------------------------------

def export_to_excel(leads: list[Lead], output_path: str, search_meta: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()

    # ===== Sheet 1: לידים =====
    ws = wb.active
    ws.title = "לידים"
    ws.sheet_view.rightToLeft = True

    headers = [
        "#", "ציון", "ציון אותיות", "שם עסק", "אתר", "URL סופי",
        "אימייל", "טלפון", "WhatsApp", "פייסבוק", "אינסטגרם",
        "כתובת", "CMS", "Copyright", "HTTPS", "מובייל",
        "בעיות עיקריות", "הערות",
        "עדיפות", "סיכום AI", "נקודת שיחה", "זווית התקפה",
        "בעיות מרכזיות", "בעיות UX", "בעיות המרה", "בעיות אמון",
        "הכנה לשיחה", "הערות AI",
    ]
    ws.append(headers)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    leads_sorted = sorted(leads, key=lambda x: x.score, reverse=True)
    for i, lead in enumerate(leads_sorted, 1):
        prep = generate_call_prep(lead)
        row = [
            i, lead.score, lead.grade, lead.business_name,
            lead.website, lead.final_url,
            lead.email, lead.phone, lead.whatsapp,
            lead.facebook, lead.instagram,
            lead.address, lead.cms, lead.last_copyright,
            "כן" if lead.has_https else "לא",
            "כן" if lead.is_mobile_friendly else "לא",
            " | ".join(lead.issues[:5]),
            lead.notes or lead.error,
            lead.priority_level,
            lead.ai_summary,
            lead.best_talking_point,
            lead.suggested_angle,
            " | ".join(lead.main_problems),
            " | ".join(lead.ux_issues),
            " | ".join(lead.conversion_issues),
            " | ".join(lead.trust_issues),
            prep,
            lead.ai_notes,
        ]
        ws.append(row)

    # רוחב עמודות
    widths = [
        4, 7, 8, 28, 30, 30, 26, 16, 14, 18, 18, 30, 12, 11, 7, 8, 40, 18,
        10, 36, 28, 24, 32, 28, 28, 28, 50, 22,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # פונט ועיצוב גוף
    body_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.font = body_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row[0].row].height = 30

    # צבעים לציונים: F=אדום, D=כתום, C=צהוב, B=ירוק בהיר, A=ירוק
    grade_colors = {
        "F": "C00000", "D": "ED7D31", "C": "FFC000",
        "B": "92D050", "A": "00B050",
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        grade_cell = row[2]
        score_cell = row[1]
        color = grade_colors.get(grade_cell.value, "FFFFFF")
        grade_cell.fill = PatternFill("solid", start_color=color)
        grade_cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        grade_cell.alignment = Alignment(horizontal="center", vertical="center")
        score_cell.font = Font(name="Arial", size=11, bold=True)
        score_cell.alignment = Alignment(horizontal="center", vertical="center")

    # קישורים לחיצים (טקסט בלבד ב-Excel לעיתים לא נפתח, במיוחד ב-RTL / Mac)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    link_wrap = Alignment(vertical="center", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (5, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            href = _url_for_href(val) if val else None
            if href:
                cell.hyperlink = href
                cell.font = link_font
                cell.alignment = link_wrap
        em = ws.cell(row=row_idx, column=7).value
        if em and isinstance(em, str) and "@" in em.strip():
            c = ws.cell(row=row_idx, column=7)
            c.hyperlink = "mailto:" + em.strip()
            c.font = link_font
            c.alignment = link_wrap
        ph = ws.cell(row=row_idx, column=8).value
        tel = _tel_href(ph) if ph else None
        if tel:
            c = ws.cell(row=row_idx, column=8)
            c.hyperlink = tel
            c.font = link_font
            c.alignment = link_wrap

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ===== Sheet 2: סיכום =====
    ws2 = wb.create_sheet("סיכום")
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22

    title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    bold = Font(name="Arial", size=11, bold=True)

    ws2["A1"] = "סיכום סריקת לידים"
    ws2["A1"].font = title_font
    ws2["A3"] = "תאריך סריקה:"; ws2["B3"] = search_meta.get("date", "")
    ws2["A4"] = "עיר:"; ws2["B4"] = search_meta.get("city", "")
    ws2["A5"] = "סוג עסק:"; ws2["B5"] = search_meta.get("type", "")
    ws2["A6"] = "סה\"כ לידים שנסרקו:"; ws2["B6"] = len(leads)
    ws2["A7"] = "לידים עם פרטי קשר:"; ws2["B7"] = sum(1 for l in leads if l.email or l.phone)

    ws2["A9"] = "פילוח לפי ציון"
    ws2["A9"].font = title_font
    ws2["A10"] = "ציון"; ws2["B10"] = "כמות"
    ws2["A10"].font = bold; ws2["B10"].font = bold
    grade_explain = {
        "F": "ישן מאוד (80+) - הזדמנות זהב",
        "D": "ישן (60-79) - מועמדים מצוינים",
        "C": "בינוני (40-59) - אפשר לפנות",
        "B": "טוב (20-39) - פחות מעניין",
        "A": "מעולה (<20) - דלגי",
    }
    for i, (g, label) in enumerate(grade_explain.items()):
        row = 11 + i
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=f'=COUNTIF(לידים!C:C,"{g}")')
        ws2.cell(row=row, column=2).font = bold

    # ===== Sheet 3: תבניות פנייה =====
    ws3 = wb.create_sheet("תבניות פנייה")
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions["A"].width = 100

    templates = [
        ("📱 הודעת WhatsApp ראשונית (קצרה)",
         "היי [שם], אני [שמך]. ראיתי את האתר של [שם העסק] ויש לי כמה רעיונות "
         "מהירים איך להפוך אותו למזמין יותר ולהוריד את אחוז הנטישה. זמינה ל-3 דקות "
         "שיחה השבוע? אני בונה אתרים בעזרת AI במחיר מאוד נגיש."),
        ("📧 מייל ראשון (פנייה קרה)",
         "שלום [שם],\n\n"
         "שמי [שמך], אני בונה אתרים מודרניים לעסקים מקומיים. "
         "נכנסתי לאתר של [שם העסק] והבחנתי במספר נקודות שמורידות לך לידים:\n\n"
         "• [בעיה 1 מהדוח]\n"
         "• [בעיה 2 מהדוח]\n"
         "• [בעיה 3 מהדוח]\n\n"
         "הכנתי דוגמה קצרה איך האתר יכול להיראות אחרי שדרוג - אשמח לשלוח אם זה רלוונטי. "
         "התהליך לוקח לי 5-7 ימים והמחיר מתחיל מ-1,500 ₪.\n\n"
         "מה דעתך?\n"
         "[שמך] | [טלפון] | [אתר]"),
        ("📞 שיחת טלפון - פתיח (15 שניות)",
         "שלום [שם], אני [שמך]. אני בונה אתרים לעסקים קטנים, "
         "וראיתי את האתר של [שם העסק]. שניה אחת אם אפשר - "
         "האתר היום לא מוצג נכון בנייד וזה אומר שאתה כנראה מאבד 60% מהפניות. "
         "יש לך 2 דקות שאסביר לך מה הייתי משדרג?"),
        ("💬 שיחה לאחר 'אין לי תקציב'",
         "אני שומע אותך. רוב העסקים שאני עובד איתם חשבו ככה לפני שראו עלות מול תועלת. "
         "אתר חדש שעולה 1,500 ₪ ומביא פנייה אחת נוספת בחודש כבר השתלם. "
         "מה התקציב שכן יהיה לך נוח? אני יכול לבנות תוכנית מדורגת."),
        ("✅ הצעת מחיר מסודרת (מבנה)",
         "דף הבית: עיצוב חדש, מותאם נייד, כפתור 'התקשר'\n"
         "דפי שירותים: עד 5 דפים\n"
         "טופס יצירת קשר עם הגעה לוואטסאפ שלך\n"
         "אופטימיזציה לגוגל בסיסית (SEO)\n"
         "עיצוב לוגו אם אין\n"
         "אחסון לשנה + דומיין\n"
         "----\n"
         "מחיר: 1,500-2,500 ₪ | זמן: 5-7 ימי עסקים | תשלום: 50% התחלה, 50% סיום"),
    ]
    row = 1
    for title, body in templates:
        ws3.cell(row=row, column=1, value=title).font = Font(
            name="Arial", size=13, bold=True, color="1F4E78")
        ws3.cell(row=row, column=1).fill = PatternFill("solid", start_color="DEEBF7")
        row += 1
        ws3.cell(row=row, column=1, value=body).alignment = Alignment(
            wrap_text=True, vertical="top")
        ws3.cell(row=row, column=1).font = Font(name="Arial", size=11)
        ws3.row_dimensions[row].height = max(80, body.count("\n") * 18 + 40)
        row += 2

    # ===== Sheet 4: מתודולוגיה =====
    ws4 = wb.create_sheet("מתודולוגיה")
    ws4.sheet_view.rightToLeft = True
    ws4.column_dimensions["A"].width = 50
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 60
    ws4["A1"] = "כיצד מחושב ציון ההתיישנות"
    ws4["A1"].font = title_font
    ws4["A3"] = "בעיה"; ws4["B3"] = "נקודות"; ws4["C3"] = "מה זה אומר"
    for c in [ws4["A3"], ws4["B3"], ws4["C3"]]:
        c.font = bold
        c.fill = PatternFill("solid", start_color="DEEBF7")
    rules = [
        ("לא מותאם למובייל (אין viewport)", 25,
         "האתר לא מתאים את עצמו לטלפון - 70% מהגלישה היום במובייל"),
        ("Flash בשימוש", 30, "טכנולוגיה שדפדפנים כבר לא תומכים בה"),
        ("עיצוב מבוסס טבלאות", 20, "טכניקת HTML מ-1998, האתר נראה כמו מ-2005"),
        ("ללא HTTPS", 15, "Google מסמן את האתר כ'לא מאובטח' לגולשים"),
        ("Copyright ישן (3+ שנים)", 15, "סימן ברור שלא מתחזקים את האתר"),
        ("'best viewed in IE' / hit counter", 15, "סימני אתר מ-2003"),
        ("jQuery ישן (גרסה 1-2)", 10, "ספריית JavaScript מיושנת"),
        ("תגיות HTML4 ישנות", 8, "<font>, <center>, <marquee>"),
        ("ללא Open Graph", 8, "האתר לא מוצג יפה בשיתוף בפייסבוק/ווצאפ"),
        ("טעינה איטית (5+ שניות)", 10, "גולשים נוטשים אחרי 3 שניות"),
        ("ללא Schema.org", 5, "Google לא מבין את התוכן לתוצאות עשירות"),
        ("ללא favicon", 4, "סימן אופייני לאתרים חובבניים"),
        ("HTML קטן מאוד (<5KB)", 5, "כמעט בטוח אתר סטטי ישן"),
    ]
    for i, (issue, pts, expl) in enumerate(rules):
        ws4.cell(row=4 + i, column=1, value=issue)
        ws4.cell(row=4 + i, column=2, value=pts).alignment = Alignment(horizontal="center")
        ws4.cell(row=4 + i, column=3, value=expl)

    ws4[f"A{4 + len(rules) + 2}"] = "סולם ציונים:"
    ws4[f"A{4 + len(rules) + 2}"].font = bold
    grade_scale = [
        ("F (80+)", "ישן מאוד - הזדמנות זהב, פני ראשונים", "C00000"),
        ("D (60-79)", "ישן - מועמדים מצוינים", "ED7D31"),
        ("C (40-59)", "בינוני - אפשר לפנות עם הצעה ממוקדת", "FFC000"),
        ("B (20-39)", "טוב יחסית - פחות דחוף", "92D050"),
        ("A (<20)", "מעולה - דלגי, אין מה לשפר", "00B050"),
    ]
    base = 4 + len(rules) + 3
    for i, (g, expl, color) in enumerate(grade_scale):
        ws4.cell(row=base + i, column=1, value=g).fill = PatternFill("solid", start_color=color)
        ws4.cell(row=base + i, column=1).font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws4.cell(row=base + i, column=3, value=expl)

    wb.save(output_path)
    print(f"  ✓ נשמר ל: {output_path}")


def export_to_html(leads: list[Lead], output_path: str, search_meta: dict) -> str:
    """
    דשבורד מקומי: בחירת ליד, תצוגת אתר (iframe), סיכום AI, הכנה לשיחה, העתקת נקודת מפתח.
    ללא שליחת הודעות — רק עזר למפעיל האנושי.
    """
    leads_sorted = sorted(leads, key=lambda x: x.score, reverse=True)
    html_path = str(Path(output_path).with_suffix(".html"))
    payload = [lead_to_dashboard_dict(l, html_path) for l in leads_sorted]
    json_safe = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")

    city = escape(search_meta.get("city", ""))
    typ = escape(search_meta.get("type", ""))
    when = escape(search_meta.get("date", ""))

    doc = f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>הכנה לשיחות — {city}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: system-ui, "Segoe UI", Arial, sans-serif; margin: 0; background: #eef1f5; color: #1a1a1a; }}
  .banner {{
    background: #1f4e78; color: #fff; padding: 10px 16px; font-size: 0.9rem;
    border-bottom: 3px solid #ffc000;
  }}
  .banner strong {{ color: #ffc000; }}
  .topmeta {{ padding: 12px 16px; color: #444; font-size: 0.88rem; }}
  .app {{ display: flex; min-height: calc(100vh - 52px); }}
  #sidebar {{
    width: 300px; min-width: 260px; background: #fff; border-left: 1px solid #ddd;
    overflow-y: auto; max-height: calc(100vh - 52px);
  }}
  .sidebar-head {{ padding: 12px; font-weight: 700; color: #1f4e78; border-bottom: 1px solid #eee; }}
  .lead-btn {{
    display: block; width: 100%; text-align: right; padding: 10px 12px;
    border: none; border-bottom: 1px solid #f0f0f0; background: #fff; cursor: pointer;
    font-size: 0.88rem; line-height: 1.35;
  }}
  .lead-btn:hover {{ background: #f5f8fc; }}
  .lead-btn.active {{ background: #deebf7; border-right: 4px solid #1f4e78; }}
  .badge {{ display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 0.72rem; margin-inline-start: 6px; vertical-align: middle; }}
  .g-F {{ background: #c00000; color: #fff; }}
  .g-D {{ background: #ed7d31; color: #fff; }}
  .g-C {{ background: #ffc000; }}
  .g-B {{ background: #92d050; }}
  .g-A {{ background: #00b050; color: #fff; }}
  .pr-high {{ background: #fde7e7; color: #a40000; }}
  .pr-medium {{ background: #fff8e6; color: #856404; }}
  .pr-low {{ background: #e8f5e9; color: #2e7d32; }}
  #main {{ flex: 1; padding: 16px; overflow-y: auto; max-height: calc(100vh - 52px); }}
  #empty-state {{ color: #666; padding: 40px; text-align: center; }}
  #detail.hidden {{ display: none; }}
  .detail-head {{ display: flex; flex-wrap: wrap; align-items: baseline; gap: 12px; margin-bottom: 12px; }}
  .detail-head h2 {{ margin: 0; font-size: 1.35rem; color: #1f4e78; }}
  .score-pill {{ font-size: 1rem; font-weight: 700; padding: 4px 10px; border-radius: 8px; background: #fff; border: 1px solid #ccc; }}
  .preview-row {{ display: grid; grid-template-columns: 220px 1fr; gap: 12px; margin-bottom: 16px; }}
  @media (max-width: 900px) {{ .preview-row {{ grid-template-columns: 1fr; }} }}
  .shot-wrap {{
    background: #fff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden;
    min-height: 160px; display: flex; align-items: center; justify-content: center;
  }}
  .shot-wrap img {{ max-width: 100%; height: auto; display: block; }}
  .shot-placeholder {{ color: #888; font-size: 0.85rem; padding: 12px; text-align: center; }}
  .frame-wrap {{
    background: #fff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden;
    min-height: 320px;
  }}
  iframe {{ width: 100%; height: 420px; border: none; }}
  .frame-note {{ font-size: 0.8rem; color: #666; padding: 8px; background: #fafafa; border-top: 1px solid #eee; }}
  .card {{
    background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 14px 16px; margin-bottom: 12px;
  }}
  .card h3 {{ margin: 0 0 8px; font-size: 0.95rem; color: #1f4e78; }}
  .callprep {{
    background: linear-gradient(135deg, #1f4e78 0%, #2d6a9f 100%); color: #fff;
    border: none; padding: 18px; border-radius: 10px;
  }}
  .callprep h3 {{ color: #ffc000; }}
  .callprep pre {{
    white-space: pre-wrap; font-family: inherit; font-size: 0.95rem; line-height: 1.5; margin: 0 0 12px;
  }}
  .btn-row {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
  button.primary {{
    background: #ffc000; color: #1a1a1a; border: none; padding: 10px 18px; border-radius: 8px;
    font-weight: 700; cursor: pointer; font-size: 0.95rem;
  }}
  button.primary:hover {{ filter: brightness(0.95); }}
  button.ghost {{
    background: rgba(255,255,255,.2); color: #fff; border: 1px solid rgba(255,255,255,.5);
    padding: 8px 14px; border-radius: 8px; cursor: pointer;
  }}
  .copy-toast {{ font-size: 0.85rem; margin-inline-start: 8px; color: #ffc000; }}
  ul.compact {{ margin: 0; padding-right: 18px; }}
  ul.compact li {{ margin-bottom: 4px; }}
  .muted {{ color: #666; font-size: 0.85rem; }}
  a {{ color: #0563c1; }}
</style>
</head>
<body>
<div class="banner">
  <strong>אנושי בלבד:</strong> הכלי לא שולח הודעות, לא מפעיל בוטים ולא יוצר ספאם — רק מכין אותך לשיחה או להודעה שאת כותבת בעצמך.
</div>
<div class="topmeta">נוצר: {when} · {city} · {typ} · <span id="nleads"></span> לידים · מיון ברשימה: לחצי על שם העסק</div>
<div class="app">
  <aside id="sidebar">
    <div class="sidebar-head">לידים (לפי ציון)</div>
    <div id="list"></div>
  </aside>
  <main id="main">
    <div id="empty-state">← בחרי ליד מהרשימה כדי לראות סיכום AI והכנה לשיחה</div>
    <div id="detail" class="hidden">
      <div class="detail-head">
        <h2 id="d-name"></h2>
        <span class="score-pill">ציון: <span id="d-score"></span> · <span id="d-grade"></span></span>
        <span id="d-pri" class="badge"></span>
        <a id="d-open" href="#" target="_blank" rel="noopener">פתחי אתר בלשונית חדשה</a>
      </div>
      <div class="preview-row">
        <div class="shot-wrap" id="d-shot-wrap">
          <span class="shot-placeholder" id="d-shot-ph">אין צילום מסך</span>
          <img id="d-shot" alt="" style="display:none;" />
        </div>
        <div>
          <div class="frame-wrap">
            <iframe id="d-frame" title="תצוגה מקדימה" sandbox="allow-same-origin allow-scripts allow-popups allow-popups-to-escape-sandbox allow-forms"></iframe>
          </div>
          <div class="frame-note">אם ה־iframe ריק — האתר חוסם הטמעה. השתמשי בקישור «פתחי אתר».</div>
        </div>
      </div>
      <div class="card">
        <h3>סיכום AI</h3>
        <p id="d-summary" class="muted"></p>
        <p id="d-ainote" class="muted" style="color:#a40000;"></p>
      </div>
      <div class="card">
        <h3>בעיות מרכזיות</h3>
        <ul class="compact" id="d-mainp"></ul>
      </div>
      <div class="card">
        <h3>UX / אמון / המרה</h3>
        <p><strong>UX:</strong></p><ul class="compact" id="d-ux"></ul>
        <p><strong>אמון:</strong></p><ul class="compact" id="d-trust"></ul>
        <p><strong>המרה:</strong></p><ul class="compact" id="d-conv"></ul>
      </div>
      <div class="callprep">
        <h3>הכנה לשיחה</h3>
        <pre id="d-prep"></pre>
        <div class="btn-row">
          <button type="button" class="primary" id="btn-copy-tp">העתקת נקודת שיחה</button>
          <button type="button" class="ghost" id="btn-copy-prep">העתקת כל תיבת ההכנה</button>
          <span class="copy-toast" id="copy-toast"></span>
        </div>
      </div>
    </div>
  </main>
</div>
<script type="application/json" id="leads-data">{json_safe}</script>
<script>
(function() {{
  const LEADS = JSON.parse(document.getElementById("leads-data").textContent);
  document.getElementById("nleads").textContent = LEADS.length;
  const listEl = document.getElementById("list");
  const emptyEl = document.getElementById("empty-state");
  const detailEl = document.getElementById("detail");
  let activeIdx = -1;

  function priClass(p) {{
    if (p === "high") return "pr-high";
    if (p === "low") return "pr-low";
    return "pr-medium";
  }}
  function priLabel(p) {{
    if (p === "high") return "עדיפות גבוהה";
    if (p === "low") return "עדיפות נמוכה";
    return "עדיפות בינונית";
  }}

  LEADS.forEach((L, i) => {{
    const b = document.createElement("button");
    b.type = "button";
    b.className = "lead-btn";
    b.dataset.idx = i;
    const g = L.grade || "?";
    b.innerHTML = (L.business_name || "ללא שם") +
      '<span class="badge g-' + g + '">' + g + "</span>" +
      '<span class="badge ' + priClass(L.priority_level) + '">' + priLabel(L.priority_level) + "</span>";
    b.addEventListener("click", () => selectLead(i));
    listEl.appendChild(b);
  }});

  function ulFill(el, items) {{
    el.innerHTML = "";
    (items || []).forEach(t => {{
      const li = document.createElement("li");
      li.textContent = t;
      el.appendChild(li);
    }});
    if (!el.children.length) {{
      const li = document.createElement("li");
      li.className = "muted";
      li.textContent = "—";
      el.appendChild(li);
    }}
  }}

  function selectLead(i) {{
    activeIdx = i;
    const L = LEADS[i];
    document.querySelectorAll(".lead-btn").forEach((x, j) => {{
      x.classList.toggle("active", j === i);
    }});
    emptyEl.style.display = "none";
    detailEl.classList.remove("hidden");

    document.getElementById("d-name").textContent = L.business_name || "";
    document.getElementById("d-score").textContent = L.score;
    document.getElementById("d-grade").textContent = L.grade || "";
    const pr = document.getElementById("d-pri");
    pr.textContent = priLabel(L.priority_level);
    pr.className = "badge " + priClass(L.priority_level);

    const url = L.final_url || L.website || "#";
    const openA = document.getElementById("d-open");
    openA.href = url;

    const frame = document.getElementById("d-frame");
    frame.src = url;

    const img = document.getElementById("d-shot");
    const ph = document.getElementById("d-shot-ph");
    if (L.screenshot_rel) {{
      img.src = L.screenshot_rel;
      img.style.display = "block";
      ph.style.display = "none";
      img.onerror = () => {{ img.style.display = "none"; ph.style.display = "block"; ph.textContent = "שגיאה בטעינת תמונה"; }};
    }} else {{
      img.src = "";
      img.style.display = "none";
      ph.style.display = "block";
      ph.textContent = "אין צילום (הריצי עם --screenshots + Playwright)";
    }}

    document.getElementById("d-summary").textContent = L.ai_summary || "(אין סיכום — בדקי מפתח API או --no-ai)";
    const an = document.getElementById("d-ainote");
    an.textContent = L.ai_notes || "";
    ulFill(document.getElementById("d-mainp"), L.main_problems);
    ulFill(document.getElementById("d-ux"), L.ux_issues);
    ulFill(document.getElementById("d-trust"), L.trust_issues);
    ulFill(document.getElementById("d-conv"), L.conversion_issues);
    document.getElementById("d-prep").textContent = L.call_prep || "";

    const tp = L.best_talking_point || "";
    document.getElementById("btn-copy-tp").onclick = () => copyText(tp, "נקודת השיחה הועתקה");
    document.getElementById("btn-copy-prep").onclick = () => copyText(L.call_prep || "", "ההכנה הועתקה");
  }}

  function copyText(text, msg) {{
    const toast = document.getElementById("copy-toast");
    if (!text) {{ toast.textContent = "אין מה להעתיק"; return; }}
    navigator.clipboard.writeText(text).then(() => {{
      toast.textContent = msg;
      setTimeout(() => {{ toast.textContent = ""; }}, 2500);
    }}).catch(() => {{
      toast.textContent = "העתקה נכשלה — סמני ידנית";
    }});
  }}

  if (LEADS.length) selectLead(0);
}})();
</script>
</body>
</html>"""
    Path(html_path).write_text(doc, encoding="utf-8")
    print(f"  ✓ דשבורד הכנה לשיחות: {html_path}")
    return html_path


# ----------------------------------------------------------------------------
# עיבוד ליד בודד
# ----------------------------------------------------------------------------

def _process_no_website_lead(lead: Lead, biz: dict, description: str = "") -> Lead:
    """
    מטפל בעסק שאין לו אתר. מציין את זה במפורש, נותן ציון C/50 (בינוני),
    ומריץ AI לפי שם/כתובת/סוג עסק כדי להעריך התאמה לתיאור.
    """
    lead.no_website = True
    # שומר את הקישור לדף הסושיאל/Google Maps שבו ראינו את הטלפון —
    # זה גם הקישור שאפשר לאמת איתו.
    lead.social_url = (biz.get("social") or biz.get("source_url") or "").strip()
    lead.score = 50          # ציון בעיות בינוני - אין על מה לבדוק
    lead.grade = "C"
    lead.issues = ["אין אתר אינטרנט עצמאי"]
    lead.has_https = False
    lead.is_mobile_friendly = False
    lead.cms = ""
    lead.last_copyright = ""
    lead.first_seen_year = 0
    lead.domain_age_years = 0
    lead.load_time_ms = 0
    lead.html_size_kb = 0.0
    lead.phone = _normalize_il_phone(lead.phone)
    # site_key: שם+טלפון (כי אין דומיין)
    phone_digits = re.sub(r"\D", "", (lead.phone or ""))
    name_norm = (lead.business_name or "").strip().lower()
    lead.site_key = f"nosite|{name_norm}|{phone_digits}"
    lead.last_analyzed_at = datetime.utcnow().isoformat() + "Z"

    # AI: התאמה לתיאור (גם בלי אתר)
    desc = (description or "").strip()
    if desc:
        api_key = (os.environ.get("OPENAI_API_KEY") or "").strip()
        if api_key:
            try:
                from openai import OpenAI
                client = OpenAI(api_key=api_key)
                prompt = f"""העסק: {lead.business_name}
כתובת: {lead.address or '—'}
טלפון: {lead.phone or '—'}
לעסק אין אתר אינטרנט.

התיאור של מי שאני מחפש (לקוח אידיאלי):
{desc}

תני match_score 0-100 והסבר קצר. החזירי JSON: {{"match_score":int,"match_reason":"..."}}"""
                completion = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "החזירי JSON תקני בלבד."},
                        {"role": "user", "content": prompt},
                    ],
                    response_format={"type": "json_object"},
                )
                import json as _json
                d = _json.loads(completion.choices[0].message.content or "{}")
                lead.match_score = int(d.get("match_score", 0) or 0)
                lead.match_reason = str(d.get("match_reason", "") or "")
            except Exception as e:
                print(f"  no-website match AI failed: {e}", file=sys.stderr)
                lead.match_score = 70
                lead.match_reason = "עסק בלי אתר — הערכת התאמה ידנית"
        else:
            lead.match_score = 70
            lead.match_reason = "עסק בלי אתר — אין OPENAI_API_KEY"
    else:
        # אין תיאור — לתת ציון 70 כדי לעבור את הסף של 50
        lead.match_score = 70
        lead.match_reason = "עסק בלי אתר — הזדמנות פתוחה"

    # שדות AI ריקים שלא רלוונטיים
    lead.ai_summary = (
        f"לעסק {lead.business_name} אין אתר אינטרנט. "
        "זו הזדמנות מצוינת להציע בניית אתר חדש מאפס."
    )
    lead.main_problems = ["אין אתר → אין נוכחות דיגיטלית מקצועית"]
    lead.strongest_problem = "אין אתר אינטרנט"
    lead.business_impact = "לקוחות פוטנציאליים לא מוצאים את העסק בחיפוש בגוגל. כל מתחרה עם אתר תופס את הלקוחות שלך."
    lead.opening_line = f"שלום, ראיתי שאין לך אתר ל{lead.business_name}. רציתי להבין למה אתם לא בגוגל - יש לך זמן ל-2 דקות?"
    lead.if_not_interested = "תוכל להגיד לי בכמה זה עוזר לך עכשיו? אולי בעוד חודש שווה לחזור?"
    lead.what_to_offer = "בניית אתר מקצועי עם דומיין, אחסון, ועיצוב מותאם לעסק שלך."
    lead.next_action = "call"

    # תסריט שיחה מותאם לעסק בלי אתר
    biz_name = lead.business_name or "העסק"
    biz_type = lead.search_business_type or "העסק שלך"
    lead.script_intro = (
        f"היי, מדברת [שם שלך]. אני בונה אתרים לעסקים. "
        f"ראיתי ש-{biz_name} עדיין לא ב-Google בצורה מסודרת — אין לכם אתר. "
        f"יש לך 2 דקות שאסביר לך מה זה אומר בפועל?"
    )
    lead.script_discovery = [
        f"איך לקוחות חדשים מוצאים אתכם היום? פייסבוק? המלצות? קופון?",
        f"כמה לקוחות חדשים אתם מקבלים בשבוע מגוגל?",
        f"חשבתם פעם על אתר? מה עצר אתכם?",
    ]
    lead.script_value_pitch = (
        f"תראה, הלקוחות שלך מחפשים '{biz_type}' בגוגל כל יום, "
        f"ובלי אתר הם פשוט מגיעים למתחרים שלך. "
        f"אתר זה לא הוצאה — זה תיק לקוחות חדשים שאת מאבד עכשיו."
    )
    lead.script_offer = (
        f"אני יכולה לבנות לך אתר פשוט אבל מקצועי תוך 2-3 שבועות. "
        f"אתר שיופיע בגוגל, יראה אותך מקצועי, ויביא לך פניות. "
        f"מחיר התחלתי שמתאים לעסק קטן."
    )
    lead.script_close = "מה דעתך שניפגש לעוד 15 דקות, אני אראה לך דוגמאות, ואת תחליטי?"
    lead.script_objections = {
        "אין לי תקציב": "מבינה. אגב, יש אצלי פתרון התחלתי שעולה פחות ממה שאתם מוציאים על פרסום בחודש. רוצה לשמוע?",
        "אני סומך על פייסבוק": "פייסבוק זה מצוין, אבל אנשים שמחפשים אותך בגוגל לא יראו אותך שם. שני הדברים משלימים.",
        "אין לי זמן עכשיו": "ברור. מתי יותר נח? אחזור אליך אז ב-5 דקות בלבד.",
        "אני אחשוב על זה": "בטח. אשלח לך 2-3 דוגמאות של עסקים דומים שעבדתי איתם — תראה ותחליט בלי לחץ. מה האימייל הכי טוב?",
    }
    lead.script_dos_and_donts = [
        "אל תקראי מהדף — דברי טבעי, כמו לחבר.",
        "תני להם לדבר. אחרי שאלה — שתקי.",
        "אם הם אמרו 'לא' פעמיים — סיימי בנימוס וחזרי בעוד חודש.",
        "אל תתחילי מהמחיר. תחילה תני להם להבין שהם מפסידים בלי אתר.",
    ]

    lead.priority_level = compute_priority_level(lead)
    enrich_lead_for_crm(lead)
    return lead


def process_lead(
    biz: dict,
    *,
    use_ai: bool = True,
    screenshot_dir: Optional[str] = None,
    description: str = "",
) -> Lead:
    lead = Lead(
        business_name=biz.get("name", ""),
        website=biz.get("website", ""),
        phone=biz.get("phone", ""),
        email=biz.get("email", ""),
        address=biz.get("address", ""),
    )
    # מצב "עסק בלי אתר" — מסלול מיוחד
    if biz.get("no_website") or not lead.website:
        return _process_no_website_lead(lead, biz, description)
    if not lead.website:
        lead.error = "אין URL"
        return lead
    try:
        score, issues, meta = analyze_website(lead.website)
        lead.score = score
        lead.grade = score_to_grade(score)
        lead.issues = issues
        lead.final_url = meta["final_url"]
        lead.has_https = meta["has_https"]
        lead.is_mobile_friendly = meta["is_mobile_friendly"]
        lead.last_copyright = meta["last_copyright"]
        lead.cms = meta["cms"]
        lead.first_seen_year = meta.get("first_seen_year", 0) or 0
        lead.domain_age_years = meta.get("domain_age_years", 0) or 0
        lead.load_time_ms = meta.get("load_time_ms", 0) or 0
        lead.html_size_kb = meta.get("html_size_kb", 0.0) or 0.0

        if not lead.email or not lead.phone:
            contacts = extract_contacts(lead.final_url, html=None)
            lead.email = lead.email or contacts["email"]
            lead.phone = lead.phone or contacts["phone"]
            lead.whatsapp = contacts["whatsapp"]
            lead.facebook = contacts["facebook"]
            lead.instagram = contacts["instagram"]

        # ניקוי איכות נתונים
        lead.final_url = _normalize_url_for_storage(lead.final_url or lead.website)
        lead.website = _normalize_url_for_storage(lead.website)
        lead.phone = _normalize_il_phone(lead.phone)
        lead.site_key = _site_key_from_url(lead.final_url or lead.website)
        lead.last_analyzed_at = datetime.utcnow().isoformat() + "Z"

        if use_ai:
            digest = meta.get("html_for_ai") or ""
            ai_d = analyze_with_ai(digest, meta.get("final_url") or lead.website, description=description)
            apply_ai_dict_to_lead(lead, ai_d)
        else:
            lead.ai_notes = "ניתוח AI כבוי (--no-ai)"

        lead.priority_level = compute_priority_level(lead)
        enrich_lead_for_crm(lead)

        if screenshot_dir and lead.final_url:
            slug = hashlib.md5(lead.final_url.encode("utf-8")).hexdigest()[:12]
            dest = Path(screenshot_dir) / f"shot_{slug}.png"
            if capture_homepage_screenshot(lead.final_url, dest):
                lead.screenshot_path = str(dest.resolve())
    except Exception as e:
        lead.error = str(e)
    if not (lead.priority_level or "").strip():
        lead.priority_level = compute_priority_level(lead)
    if lead.website:
        enrich_lead_for_crm(lead)
    return lead


class NoBusinessesFound(Exception):
    """לא נמצאו עסקים לחיפוש המבוקש."""


def run_pipeline(
    city: str,
    business_type: str,
    limit: int = 30,
    out_path: Optional[str] = None,
    workers: int = 5,
    export_html: bool = True,
    use_ai: bool = True,
    screenshots: bool = False,
    on_progress: Optional[Callable[[int, int, Lead], None]] = None,
    quiet: bool = False,
    skip_export: bool = False,
    description: str = "",
    exclude_domains: Optional[set[str]] = None,
    only_without_website: bool = False,
) -> dict:
    """
    מריץ חיפוש + ניתוח + ייצוא. מחזיר מילון עם leads, נתיבי קבצים ו-meta.
    on_progress(index, total, lead) — אופציונלי (לדשבורד).
    """
    out = out_path or f"leads_{city}_{business_type}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    out = re.sub(r'[<>:"/\\|?*]', '_', out)

    if not use_ai and not quiet:
        pass
    elif use_ai and not (os.environ.get("OPENAI_API_KEY") or "").strip() and not quiet:
        print(
            "  ⚠ אין OPENAI_API_KEY — ניתוח AI ידולג. "
            "צרי קובץ .env ליד find_leads.py (ראי .env.example) או export OPENAI_API_KEY=... או --no-ai",
            file=sys.stderr,
        )

    shot_dir: Optional[str] = None
    if screenshots:
        shot_dir = str((Path(out).resolve().parent / "lead_screenshots").resolve())
        Path(shot_dir).mkdir(parents=True, exist_ok=True)

    def _job(b: dict) -> Lead:
        return process_lead(b, use_ai=use_ai, screenshot_dir=shot_dir, description=description)

    businesses = find_businesses(
        city,
        business_type,
        limit,
        description=description,
        exclude_domains=exclude_domains,
        only_without_website=only_without_website,
    )
    if not businesses:
        raise NoBusinessesFound(f"לא נמצאו עסקים עבור '{business_type}' ב'{city}'")

    if not quiet:
        print(f"\nמנתח {len(businesses)} אתרים...")
    leads: list[Lead] = []
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_job, b): b for b in businesses}
        for i, fut in enumerate(as_completed(futures), 1):
            lead = fut.result()
            leads.append(lead)
            if on_progress:
                on_progress(i, len(businesses), lead)
            if not quiet:
                tag = f"[{lead.grade}/{lead.score}]" if not lead.error else "[ERR]"
                print(f"  {i}/{len(businesses)} {tag} {lead.business_name[:30]}")

    meta = {
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "city": city,
        "type": business_type,
    }
    for L in leads:
        L.search_city = city
        L.search_business_type = business_type

    # אם יש תיאור — סנן לידים לא רלוונטיים (match_score < 50) ומיין את השאר
    if (description or "").strip():
        before_n = len(leads)
        relevant = [L for L in leads if (L.match_score or 0) >= 50]
        dropped = before_n - len(relevant)
        if dropped and not quiet:
            print(f"\n🚫 סוננו {dropped} לידים עם match_score < 50 (לא רלוונטיים לתיאור)")
        leads = relevant
        leads.sort(key=lambda L: (L.match_score or 0, L.opportunity_score or 0), reverse=True)
        if not quiet:
            print(f"\n📊 מויין לפי התאמה לתיאור:")
            for L in leads[:5]:
                ms = L.match_score or 0
                print(f"  • {L.business_name[:30]} — match_score={ms}")

    html_abs: Optional[str] = None
    out_abs: Optional[str] = None
    if not skip_export:
        if not quiet:
            print(f"\nמייצא לאקסל...")
        export_to_excel(leads, out, meta)
        if export_html:
            html_abs = export_to_html(leads, out, meta)
        out_abs = str(Path(out).resolve())
    f_count = sum(1 for l in leads if l.grade == "F")
    d_count = sum(1 for l in leads if l.grade == "D")
    contact_n = sum(1 for l in leads if l.email or l.phone)
    if not quiet:
        print(f"\n✓ סיימתי. מתוך {len(leads)} לידים:")
        print(f"  • {f_count} ישנים מאוד (F) - פני אליהם ראשונים!")
        print(f"  • {d_count} ישנים (D) - מועמדים מצוינים")
        print(f"  • {contact_n} עם פרטי קשר")
        if out_abs:
            print(f"\nקובץ פלט: {out_abs}")
        if html_abs:
            print(f"דף עבודה בדפדפן (קישורים לחיצים): {html_abs}")

    return {
        "leads": leads,
        "out_xlsx": out_abs,
        "out_html": str(Path(html_abs).resolve()) if html_abs else None,
        "meta": meta,
        "stats": {
            "total": len(leads),
            "grade_f": f_count,
            "grade_d": d_count,
            "with_contact": contact_n,
        },
    }


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    _load_env_file()
    p = argparse.ArgumentParser(
        description="מחפש לידים + עוזר הכנה לשיחה (אנושי בלבד — ללא שליחת הודעות)"
    )
    p.add_argument("--city", required=True, help="עיר/אזור (למשל: 'תל אביב')")
    p.add_argument("--type", required=True, dest="business_type", help="סוג עסק")
    p.add_argument("--limit", type=int, default=30, help="כמה לידים לחפש (ברירת מחדל 30)")
    p.add_argument("--out", default=None, help="נתיב פלט אקסל")
    p.add_argument("--workers", type=int, default=5, help="מספר חוטים לסריקה")
    p.add_argument("--no-html", action="store_true", help="ללא ייצוא דף HTML (רק אקסל)")
    p.add_argument("--no-ai", action="store_true", help="ללא ניתוח AI (חוסך עלות API)")
    p.add_argument(
        "--screenshots",
        action="store_true",
        help="צילומי דף הבית (דורש: pip install playwright && playwright install chromium)",
    )
    args = p.parse_args()

    try:
        run_pipeline(
            city=args.city,
            business_type=args.business_type,
            limit=args.limit,
            out_path=args.out,
            workers=args.workers,
            export_html=not args.no_html,
            use_ai=not args.no_ai,
            screenshots=args.screenshots,
        )
    except NoBusinessesFound as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
